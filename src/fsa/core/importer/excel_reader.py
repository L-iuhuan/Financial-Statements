"""Excel/CSV 文件读取器: 支持 .xlsx/.xlsm (openpyxl)、.xls (pandas+xlrd) 与 .csv。

CSV 编码自动探测 UTF-8(含 BOM) 与 GBK。

统一转换为 RawSheetData，并为导入框架提供通用能力:
- 自动定位表头行（兼容标题行、合并单元格、无"项目"列的场景）
- 捕获连续的多行表头（header_rows），供 SCE 等多级表头使用
- 数据行以列标题为键，额外包含 "_row" 键（源文件行号，从 1 开始）
"""

from __future__ import annotations

import csv
import os
import re
import threading
from collections.abc import Callable, Sequence
from dataclasses import dataclass, field
from pathlib import Path
from types import ModuleType
from typing import cast
from zipfile import BadZipFile

import openpyxl
from loguru import logger
from openpyxl.worksheet.worksheet import Worksheet

from fsa.core.exceptions import FSAError

_NAME_HEADER_CANDIDATES = ("项目", "项目名称", "科目", "科目名称")
_AMOUNT_HEADER_KEYWORDS = (
    "期末余额",
    "年初余额",
    "本期金额",
    "上期金额",
    "期末数",
    "年初数",
    "本期数",
    "上期数",
    "金额",
    "余额",
    "发生额",
)
_MAX_SCAN_ROWS = 15
_MAX_HEADER_LAYERS = 4
# 超过该字节数的 .xlsx/.xlsm 优先用 read_only 流式读取, 降低大文件内存峰值
_LARGE_FILE_THRESHOLD = 5 * 1024 * 1024

try:
    import xlrd

    _XLRD_ERRORS: tuple[type[Exception], ...] = (xlrd.biffh.XLRDError,)
except ImportError:
    _XLRD_ERRORS = ()

# 常规读取失败类型: 缺依赖 (ImportError, 如 .xls 缺 xlrd/pandas)、
# 加密/损坏文件 (BadZipFile/XLRDError) 等统一回退到 Excel COM 读取
_NATIVE_READ_ERRORS = (
    BadZipFile,
    OSError,
    ValueError,
    KeyError,
    TypeError,
    ImportError,
) + _XLRD_ERRORS


@dataclass
class RawSheetData:
    """一个工作表的原始数据。

    Attributes:
        name: 工作表名称
        headers: 主表头行（第一层表头）
        header_rows: 连续的多层表头（至少 1 行），供矩阵式报表使用
        rows: 数据行列表，每行是 dict，键为 headers 中的列标题，
              额外包含 "_row" 键（源文件行号，从 1 开始）
    """

    name: str
    headers: list[str] = field(default_factory=list)
    header_rows: list[list[str]] = field(default_factory=list)
    rows: list[dict[str, object]] = field(default_factory=list)


def read_excel(file_path: str, use_com: bool = False) -> dict[str, RawSheetData]:
    """读取 Excel 文件，返回所有工作表的原始数据。

    Args:
        file_path: Excel 文件路径（.xlsx 或 .xls）
        use_com: 强制使用 Excel COM 读取（默认为 False，常规读取失败时自动回退）

    Returns:
        字典，键为工作表名称，值为 RawSheetData

    Raises:
        FileNotFoundError: 文件不存在
        FSAError: 常规读取与 Excel COM 读取均失败
    """
    path = str(file_path)
    if use_com:
        return read_excel_com(path)

    try:
        return _read_native(path)
    except FileNotFoundError:
        raise
    except _NATIVE_READ_ERRORS as error:
        logger.warning(f"常规方式读取失败（{error}），尝试用 Excel COM 打开: {path}")
        try:
            return read_excel_com(path)
        except FSAError as com_error:
            raise FSAError(f"文件「{path}」常规解析失败（{error}），Excel COM 打开也失败: {com_error}") from error


def _read_native(path: str) -> dict[str, RawSheetData]:
    """用 openpyxl / pandas+xlrd 常规读取 Excel 文件。

    .xlsx/.xlsm 大文件 (>5MB) 优先 read_only 流式读取; 流式读取失败
    (某些加密/兼容性文件) 时回落普通模式, 再由 read_excel 的 COM 回退兜底。
    """
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xls":
        return _read_xls(path)
    # .xlsx / .xlsm 均由 openpyxl 读取 (含宏文件的数据部分)
    try:
        size = os.path.getsize(path)
    except OSError as e:
        raise FileNotFoundError(f"无法打开文件「{path}」: {e}") from e
    if size >= _LARGE_FILE_THRESHOLD:
        try:
            return _read_openpyxl(path, read_only=True)
        except (OSError, ValueError, KeyError, TypeError) as e:
            logger.warning(f"大文件流式读取失败, 回落普通模式: {path} ({e})")
    return _read_openpyxl(path, read_only=False)


def _read_openpyxl(path: str, read_only: bool) -> dict[str, RawSheetData]:
    """openpyxl 读取 .xlsx/.xlsm, 支持 read_only 流式模式。"""
    logger.info(f"正在读取 Excel 文件: {path} (read_only={read_only})")
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=read_only)
    except FileNotFoundError:
        raise
    except OSError as e:
        raise FileNotFoundError(f"无法打开文件「{path}」: {e}") from e

    result: dict[str, RawSheetData] = {}
    for sheet_name in wb.sheetnames:
        matrix = _sheet_to_matrix(wb[sheet_name])
        raw = _matrix_to_raw(sheet_name, matrix)
        result[sheet_name] = raw
        logger.debug(f"  工作表「{sheet_name}」: {len(raw.rows)} 行数据")

    wb.close()
    logger.info(f"读取完成，共 {len(result)} 个工作表")
    return result


def _sheet_to_matrix(worksheet: Worksheet) -> list[list[object]]:
    """将 openpyxl 工作表转换为二维矩阵（行优先）。

    使用 iter_rows(values_only=True) 批量读取, 避免逐单元格 cell() 调用;
    大工作表下性能可提升一个数量级。
    """
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _read_xls(path: str) -> dict[str, RawSheetData]:
    """通过 pandas + xlrd 读取旧版 .xls 文件。"""
    try:
        import pandas as pd
    except ImportError as e:
        raise ImportError("读取 .xls 需要安装 pandas 与 xlrd") from e

    logger.info(f"正在读取 .xls 文件: {path}")
    try:
        sheets = pd.read_excel(path, sheet_name=None, header=None, dtype=object, engine="xlrd")
    except OSError as e:
        raise FileNotFoundError(f"无法打开文件「{path}」: {e}") from e

    result: dict[str, RawSheetData] = {}
    for sheet_name, frame in sheets.items():
        matrix = [[cell for cell in row] for row in frame.values.tolist()]
        raw = _matrix_to_raw(sheet_name, matrix)
        result[sheet_name] = raw
        logger.debug(f"  工作表「{sheet_name}」: {len(raw.rows)} 行数据")

    logger.info(f".xls 读取完成，共 {len(result)} 个工作表")
    return result


def _read_csv(path: str) -> dict[str, RawSheetData]:
    """读取 CSV 文件 (UTF-8/GBK 编码探测), 转换为单工作表 RawSheetData。"""
    text = _read_csv_text(path)
    rows = list(csv.reader(text.splitlines(), skipinitialspace=True))
    if not rows:
        raise FSAError(f"CSV 文件「{path}」为空")
    matrix: list[list[object]] = [[cell if cell != "" else None for cell in row] for row in rows]
    # 空值统一为 None, 与 Excel 读取路径行为一致
    sheet_name = Path(path).stem
    raw = _matrix_to_raw(sheet_name, matrix)
    logger.info(f"CSV 读取完成: 工作表「{sheet_name}」, {len(raw.rows)} 行数据")
    return {sheet_name: raw}


def _read_csv_text(path: str) -> str:
    """按 UTF-8(含 BOM)/GBK 顺序探测 CSV 文本编码。"""
    try:
        data = Path(path).read_bytes()
    except OSError as e:
        raise FileNotFoundError(f"无法打开文件「{path}」: {e}") from e
    for encoding in ("utf-8-sig", "gbk", "utf-8"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise FSAError(f"CSV 文件「{path}」编码无法识别 (已尝试 UTF-8/GBK)")


def _matrix_to_raw(
    name: str, matrix: list[list[object]], row_offset: int = 0
) -> RawSheetData:
    """从二维矩阵构建 RawSheetData（openpyxl 与 pandas 路径共用）。

    row_offset: 矩阵第 0 行在源文件中的实际行号 - 1（COM UsedRange 偏移），
        用于把 `_row` 还原为源文件 1-based 行号。
    """
    if not matrix:
        return RawSheetData(name=name, headers=[], header_rows=[], rows=[])
    header_idx = _find_header_row(matrix)
    header_rows = _capture_header_rows(matrix, header_idx)
    headers = _uniquify(_merged_headers(header_rows))
    rows = _build_rows(matrix, headers, header_idx + len(header_rows), row_offset)
    return RawSheetData(name=name, headers=headers, header_rows=header_rows, rows=rows)


def _merged_headers(header_rows: list[list[str]]) -> list[str]:
    """把多层表头按列纵向合并为单层列名。

    例如:
        行1: 项目 | 期末余额 | 年初余额
        行2:      | 2024年   | 2023年
    合并为:  项目 | 期末余额2024年 | 年初余额2023年

    这样后续"候选词包含匹配"仍能命中"期末余额/年初余额"等标准列名,
    同时合并单元格留下的空位 (仅左上角有值) 也能取到下层标签。
    只有一层表头时与原有 _to_header_list 行为一致。
    """
    if not header_rows:
        return []
    column_count = max((len(row) for row in header_rows), default=0)
    headers: list[str] = []
    for col_idx in range(column_count):
        parts: list[str] = []
        for row in header_rows:
            if col_idx >= len(row):
                continue
            label = str(row[col_idx]).strip() if row[col_idx] is not None else ""
            if label:
                parts.append(label)
        headers.append("".join(parts) if parts else f"列{col_idx + 1}")
    return headers


def _find_header_row(matrix: list[list[object]]) -> int:
    """在矩阵前若干行中定位表头行。

    优先返回含"项目/科目"类单元格的行；若不存在（如资产负债表的
    "资 产 | 行次 | 期末余额 | 年初余额"表头），则回退到含金额列
    关键词的行；仍找不到时返回第 1 行。
    """
    for row_idx in range(min(_MAX_SCAN_ROWS, len(matrix))):
        normalized = [_normalize_cell(cell) for cell in matrix[row_idx]]
        if any(value in _NAME_HEADER_CANDIDATES for value in normalized if value):
            return row_idx

    for row_idx in range(min(_MAX_SCAN_ROWS, len(matrix))):
        joined = "".join(_normalize_cell(cell) for cell in matrix[row_idx])
        if any(keyword in joined for keyword in _AMOUNT_HEADER_KEYWORDS):
            return row_idx

    return 0


def _capture_header_rows(matrix: list[list[object]], header_idx: int) -> list[list[str]]:
    """捕获从表头行开始的连续多层表头。

    子表头行判定: 首列为空且行内存在非空标签（如权益变动表的
    "股本/资本公积"层与"优先股/永续债"层）。遇到首列非空的行
    （数据行）即停止。
    """
    header_rows = [_to_header_list(matrix[header_idx], fill_empty=False)]
    for offset in range(1, _MAX_HEADER_LAYERS):
        row_idx = header_idx + offset
        if row_idx >= len(matrix):
            break
        cells = matrix[row_idx]
        if _contains_number(cells):
            break
        first = _normalize_cell(cells[0]) if cells else ""
        if first:
            break
        if not any(_normalize_cell(cell) for cell in cells):
            break
        header_rows.append(_to_header_list(cells, fill_empty=False))
    return header_rows


def _contains_number(cells: list[object]) -> bool:
    """判断一行是否包含数值（数据行特征，用于停止多层表头捕获）。"""
    return any(isinstance(cell, (int, float)) and not isinstance(cell, bool) for cell in cells)


def _build_rows(
    matrix: list[list[object]],
    headers: list[str],
    start_idx: int,
    row_offset: int = 0,
) -> list[dict[str, object]]:
    """从表头之后开始，将矩阵行转换为以列标题为键的字典列表。

    row_offset 为矩阵相对源文件的行偏移 (COM UsedRange 起始行 - 1)。
    """
    rows: list[dict[str, object]] = []
    for row_idx in range(start_idx, len(matrix)):
        cells = matrix[row_idx]
        row_data: dict[str, object] = {"_row": row_idx + 1 + row_offset}
        has_value = False
        for col_idx, header in enumerate(headers):
            value = cells[col_idx] if col_idx < len(cells) else None
            if value is not None:
                has_value = True
            row_data[header] = value
        if has_value:
            rows.append(row_data)
    return rows


def _to_header_list(cells: Sequence[object], fill_empty: bool) -> list[str]:
    """将一行单元格转换为列标题列表。

    Args:
        cells: 一行单元格
        fill_empty: 空单元格是否用「列N」占位（headers 需要，多层表头标签不需要）
    """
    headers: list[str] = []
    for col_idx, cell in enumerate(cells, 1):
        if cell is None or str(cell).strip() == "":
            headers.append(f"列{col_idx}" if fill_empty else "")
        else:
            headers.append(str(cell))
    return headers


def _uniquify(headers: list[str]) -> list[str]:
    """为重复的列标题追加序号后缀，避免按列名建 dict 时互相覆盖。

    例如资产负债表的左右两栏都叫「期末余额」，处理后变为
    「期末余额」与「期末余额#2」。
    """
    counts: dict[str, int] = {}
    result: list[str] = []
    for header in headers:
        count = counts.get(header, 0) + 1
        counts[header] = count
        result.append(header if count == 1 else f"{header}#{count}")
    return result


def _normalize_cell(value: object) -> str:
    """去除单元格文本中所有空白（含全角空格），用于表头匹配。"""
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value))


def read_excel_com(
    file_path: str,
    timeout: float | None = 60.0,
    progress_cb: Callable[[int, int], None] | None = None,
) -> dict[str, RawSheetData]:
    """通过 Excel COM（pywin32）读取 Excel 文件。

    适用于透明加密环境（如 DLP）：openpyxl/xlrd 看到的是密文，
    而 Excel 本体被加密客户端信任，可透明解密。

    COM 调用是阻塞式的, 这里放到独立 daemon 线程执行并用 join(timeout)
    施加软超时: 超时后本函数立即返回错误, 后台线程会在 Excel 响应后自行
    关闭进程 (Python 无法强杀线程, 这是 Windows COM 场景的务实取舍)。

    Args:
        file_path: Excel 文件路径（.xlsx / .xls / .xlsm）
        timeout: 超时秒数; None 表示不限时
        progress_cb: 进度回调 (已完成工作表数, 工作表总数)

    Returns:
        字典，键为工作表名称，值为 RawSheetData

    Raises:
        FSAError: 未安装 pywin32、无法启动 Excel、文件无法打开或超时
    """
    try:
        import win32com.client  # noqa: F401 - 仅探测依赖是否可用
    except ImportError as error:
        raise FSAError("未安装 pywin32，无法使用 Excel COM 读取加密文件") from error

    box: dict[str, object] = {}

    def work() -> None:
        try:
            box["data"] = _read_excel_com_sync(str(file_path), progress_cb)
        except Exception as exc:
            box["error"] = exc

    thread = threading.Thread(target=work, daemon=True, name="excel-com-reader")
    thread.start()
    thread.join(timeout)
    if thread.is_alive():
        raise FSAError("Excel COM 读取超时，请关闭占用中的 Excel 窗口后重试")
    err = box.get("error")
    if err is not None:
        if not isinstance(err, BaseException):
            raise FSAError(f"Excel COM 读取失败: {err}")
        if isinstance(err, FSAError):
            raise err
        raise FSAError(f"Excel COM 读取失败: {err}") from err
    return cast("dict[str, RawSheetData]", box["data"])


def _read_excel_com_sync(
    file_path: str,
    progress_cb: Callable[[int, int], None] | None = None,
) -> dict[str, RawSheetData]:
    """在调用线程内执行 Excel COM 读取 (不得直接调用, 经 read_excel_com 包装)。"""
    try:
        import win32com.client
    except ImportError as error:
        raise FSAError("未安装 pywin32，无法使用 Excel COM 读取加密文件") from error

    pythoncom: ModuleType | None
    co_initialized = False
    try:
        import pythoncom as _pythoncom

        pythoncom = _pythoncom
    except ImportError:
        pythoncom = None
    if pythoncom is not None:
        # pywin32 存根不含 com_error 属性, 经 getattr 取回退 OSError
        com_error = getattr(pythoncom, "com_error", OSError)
        try:
            pythoncom.CoInitialize()
            co_initialized = True
        except (com_error, OSError):
            co_initialized = False

    path = str(file_path)
    excel = None
    try:
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
        except Exception as error:
            # 防御性兜底: pywintypes.com_error 等 COM 异常类型不统一,
            # 且不同 Excel 版本差异大, 统一转为中文业务异常 (见 7.1 DLP)
            raise FSAError(f"无法启动 Excel 进程: {error}") from error
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        try:
            workbook = excel.Workbooks.Open(path, UpdateLinks=0, ReadOnly=True, AddToMru=False)
        except Exception as error:  # COM 异常类型不统一，统一转为业务异常
            raise FSAError(f"Excel 无法打开文件「{path}」: {error}") from error

        result: dict[str, RawSheetData] = {}
        try:
            sheet_count = workbook.Worksheets.Count
            if progress_cb is not None:
                progress_cb(0, sheet_count)
            for completed, sheet in enumerate(workbook.Worksheets, 1):
                used_range = sheet.UsedRange
                # UsedRange 的左上角不保证是 A1 (工作表顶部可能有空行被裁掉):
                # UsedRange.Value 返回的矩阵以 UsedRange 首行为第 0 行, 需把
                # 起始行偏移 (UsedRange.Row - 1) 传回, 否则 _row 源行号整体偏小,
                # 追溯定位会指向错误的 Excel 行 (P3)。
                row_offset = int(used_range.Row) - 1
                matrix = _com_range_to_matrix(used_range.Value)
                result[sheet.Name] = _matrix_to_raw(sheet.Name, matrix, row_offset=row_offset)
                if progress_cb is not None:
                    progress_cb(completed, sheet_count)
        finally:
            workbook.Close(SaveChanges=False)
        logger.info(f"Excel COM 读取完成，共 {len(result)} 个工作表")
        return result
    finally:
        if excel is not None:
            excel.Quit()
        if pythoncom is not None and co_initialized:
            pythoncom.CoUninitialize()


def _com_range_to_matrix(value: object) -> list[list[object]]:
    """将 Excel COM Range.Value 转换为行优先二维矩阵。"""
    if value is None:
        return []
    if not isinstance(value, (tuple, list)):
        return [[value]]
    matrix: list[list[object]] = []
    for row in value:
        row_values = list(row) if isinstance(row, (tuple, list)) else [row]
        matrix.append(list(row_values))
    return matrix
