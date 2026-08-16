"""审计底稿 Excel 导出器。

将 ValidationSummary 导出为格式化的 Excel 审计底稿，包含三个 sheet:
- 校验汇总: 统计概览和结论
- 校验明细: 每条规则的校验结果
- 科目追溯: 每个科目的来源追踪 (P3 可审计可溯源)
"""

from __future__ import annotations

import json
from datetime import datetime

from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from fsa.core.exporter._styles import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    FILL_GRAY,
    FILL_GREEN,
    FILL_HEADER,
    FILL_RED,
    FILL_YELLOW,
    FONT_HEADER,
    FONT_NORMAL,
    FONT_TITLE,
    THIN_BORDER,
)
from fsa.core.models.result import ValidationResult, ValidationSummary
from fsa.core.resources import resource_path

# PDF 来源行号编码基数 (见 importer/pdf_reader.py, D-01):
# PDF 行号 = 页码 * _PDF_ROW_BASE + 表内行号（1-based）。
# Excel 工作表最大行号为 1,048,576，恒小于该基数，
# 因此 `row >= _PDF_ROW_BASE` 恒为 PDF 来源，可安全解码为「第X页表内第N行」。
_PDF_ROW_BASE = 10_000_000

# 规则库文件路径 (支持开发模式与 PyInstaller 冻结模式)
_RULES_FILE = resource_path("cas_gouji_rule_library.json")


def _format_source_row(row: int) -> str | int:
    """将追溯行号转为可定位显示值 (P3 可审计可溯源)。

    - row <= 0: 阈值变量或未在报表中找到的科目, 无源行号, 显示空字符串
    - row >= _PDF_ROW_BASE: PDF 来源, 解码为「第X页表内第N行」
    - 其余: Excel 来源, 直接显示工作表 1-based 行号

    Args:
        row: TraceItem 的行号

    Returns:
        可定位显示值 (PDF 为中文串, Excel 为数字, 无行号为 "")
    """
    if row <= 0:
        return ""
    if row >= _PDF_ROW_BASE:
        page, table_row = divmod(row, _PDF_ROW_BASE)
        return f"第{page}页表内第{table_row}行"
    return row


def _get_rule_library_version() -> str:
    """从规则库 JSON 文件中读取版本号。"""
    try:
        with open(_RULES_FILE, encoding="utf-8") as f:
            data = json.load(f)
        return str(data.get("ruleLibrary", {}).get("version", ""))
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return ""


def _safe_text(value: str) -> str:
    """将可能被误解析为公式的字符串转为安全文本。

    当字符串以 =、+、-、@ 开头（去除首尾空白后）时，前缀单引号 ' 强制
    openpyxl 将其作为文本写入，防止公式注入（如 "=HYPERLINK(...)" 被当作
    公式执行）。

    Args:
        value: 原始字符串

    Returns:
        安全文本字符串
    """
    stripped = value.strip()
    if stripped and stripped[0] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def _is_threshold_rule(formula: str) -> bool:
    """判断公式是否为阈值/布尔型规则 (无 '==' 比较运算符)。

    阈值规则 (如 "net_profit >= 0") 和布尔规则没有左右侧对比,
    其 left_value/right_value/diff 无实际意义 (均为 0.0),
    导出时应显示 "—" 而非 0.00。

    判定标准: 公式中不包含 "==" 字符串, 即无等值比较。
    这是保守判定 — 少数包含 "==" 但实际为阈值规则的公式不会被误判,
    因为其 left_value/right_value 由引擎实际计算得出, 显示 0.00 也无害。

    Args:
        formula: 规则公式字符串

    Returns:
        True 表示阈值/布尔规则
    """
    return "==" not in formula


def _dash_for_amounts() -> str:
    """阈值规则和跳过行的金额显示占位符。"""
    return "—"


class AuditExporter:
    """审计底稿 Excel 导出器。

    接收 ValidationSummary，输出格式化的 .xlsx 文件。
    职责: 接收校验结果 -> 输出文件。禁止修改校验结果。
    """

    def export(self, summary: ValidationSummary, file_path: str) -> None:
        """将校验汇总导出为 Excel 审计底稿。

        Args:
            summary: 校验汇总结果
            file_path: 输出文件路径 (.xlsx)

        Raises:
            PermissionError: 文件被占用
            OSError: 写入失败
            RuntimeError: 内部状态异常 (如 active sheet 缺失)
        """
        wb = Workbook()

        # Sheet 1: 校验汇总
        ws_summary = wb.active
        if ws_summary is None:
            raise RuntimeError("无法获取活动工作表, 工作簿初始化异常")
        ws_summary.title = "校验汇总"
        self._write_summary_sheet(ws_summary, summary)

        # Sheet 2: 校验明细
        ws_detail = wb.create_sheet("校验明细")
        self._write_detail_sheet(ws_detail, summary)

        # Sheet 3: 科目追溯
        ws_trace = wb.create_sheet("科目追溯")
        self._write_trace_sheet(ws_trace, summary)

        # Sheet 4: 底稿说明 (源文件/哈希/规则版本/复核留痕)
        ws_info = wb.create_sheet("底稿说明")
        self._write_workpaper_info_sheet(ws_info, summary)

        wb.save(file_path)

    def _write_workpaper_info_sheet(
        self, ws: Worksheet, summary: ValidationSummary
    ) -> None:
        """写入底稿说明: 审计证据链与编制/复核留痕区。"""
        ws.merge_cells("A1:B1")
        title = ws.cell(row=1, column=1, value="审计底稿说明")
        title.font = FONT_TITLE
        title.alignment = ALIGN_CENTER

        rows: list[tuple[str, str]] = [
            ("报告期间", summary.period or "未设置"),
            ("编制时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("规则库版本", summary.rule_version or _get_rule_library_version()),
            ("源文件", "\n".join(summary.source_files) if summary.source_files else "未记录"),
            (
                "源文件 SHA256",
                "\n".join(summary.source_hashes) if summary.source_hashes else "未记录",
            ),
            (
                "源文件大小(字节)",
                "\n".join(str(size) for size in summary.source_file_sizes)
                if summary.source_file_sizes
                else "未记录",
            ),
            (
                "金额单位留痕",
                "\n".join(summary.amount_unit_notes)
                if summary.amount_unit_notes
                else "未记录",
            ),
            ("说明", "本底稿由勾稽规则引擎自动生成，公式列为规则公式文本，金额单位已统一为元。"),
            ("行号口径", "PDF 来源的原始行号为『第X页表内第N行』，N 含表头行。"),
            ("编制人", ""),
            ("复核人", ""),
            ("复核意见", ""),
        ]
        for i, (label, value) in enumerate(rows, start=2):
            label_cell = ws.cell(row=i, column=1, value=label)
            label_cell.font = Font(name="微软雅黑", size=10, bold=True)
            label_cell.alignment = ALIGN_LEFT
            label_cell.border = THIN_BORDER
            value_cell = ws.cell(row=i, column=2, value=value)
            value_cell.font = FONT_NORMAL
            value_cell.alignment = ALIGN_LEFT
            value_cell.border = THIN_BORDER
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 80

    # ── 校验汇总 sheet ──

    def _write_summary_sheet(
        self, ws: Worksheet, summary: ValidationSummary
    ) -> None:
        """写入校验汇总 sheet。"""
        ws.merge_cells("A1:B1")
        title_cell = ws.cell(row=1, column=1, value="财务报表勾稽校验审计底稿")
        title_cell.font = FONT_TITLE
        title_cell.alignment = ALIGN_CENTER

        rows: list[tuple[str, str]] = [
            ("报告期间", summary.period),
            ("校验时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("规则总数", str(summary.total)),
            ("通过", str(summary.passed)),
            ("不通过", str(summary.failed)),
            ("异常", str(summary.errored)),
            ("跳过", str(summary.skipped)),
            ("涉及报表", "、".join(rt.value for rt in summary.report_types)),
            ("规则库版本", _get_rule_library_version()),
        ]

        if summary.failed + summary.errored == 0:
            rows.append(("结论", "全部通过 ✓"))
        else:
            count = summary.failed + summary.errored
            rows.append(("结论", f"存在 {count} 项不通过或异常"))

        for i, (label, value) in enumerate(rows, start=2):
            label_cell = ws.cell(row=i, column=1, value=label)
            label_cell.font = Font(name="微软雅黑", size=10, bold=True)
            label_cell.alignment = ALIGN_RIGHT
            label_cell.border = THIN_BORDER

            value_cell = ws.cell(row=i, column=2, value=value)
            value_cell.font = FONT_NORMAL
            value_cell.alignment = ALIGN_LEFT
            value_cell.border = THIN_BORDER

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 40

    # ── 校验明细 sheet ──

    def _write_detail_sheet(
        self, ws: Worksheet, summary: ValidationSummary
    ) -> None:
        """写入校验明细 sheet。"""
        headers = [
            "规则ID", "规则名称", "分类", "校验结果",
            "左侧值", "右侧值", "差额", "容差", "公式", "说明",
        ]
        self._write_header_row(ws, headers)

        for i, result in enumerate(summary.results, start=2):
            self._write_detail_row(ws, i, result)

        # 列宽
        col_widths = [14, 28, 14, 10, 18, 18, 18, 10, 40, 40]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

    def _write_detail_row(
        self, ws: Worksheet, row: int, result: ValidationResult
    ) -> None:
        """写入校验明细的一行数据。"""
        # 确定状态文本和颜色
        if result.errored:
            status = "异常"
            fill = FILL_YELLOW
        elif result.skipped:
            status = "跳过"
            fill = FILL_GRAY
        elif result.passed:
            status = "通过"
            fill = FILL_GREEN
        else:
            status = "不通过"
            fill = FILL_RED

        # 阈值/布尔规则 (无 '==') 和跳过规则: 金额列显示 "—"
        threshold = _is_threshold_rule(result.formula)
        dash = _dash_for_amounts()
        if result.skipped or threshold:
            lv: str | float = dash
            rv: str | float = dash
            dv: str | float = dash
        else:
            lv = result.left_value
            rv = result.right_value
            dv = result.diff

        values = [
            result.rule_id,
            result.rule_name,
            result.category,
            status,
            lv,
            rv,
            dv,
            result.tolerance,
            _safe_text(result.formula),
            _safe_text(result.message),
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER

            if col == 4:  # 校验结果列
                cell.fill = fill
                cell.alignment = ALIGN_CENTER
            elif col in (5, 6, 7):  # 金额列
                if isinstance(value, str):
                    # "—" 占位符 — 居中对齐，不设数字格式
                    cell.alignment = ALIGN_CENTER
                else:
                    cell.number_format = "#,##0.00"
                    cell.alignment = ALIGN_RIGHT
            elif col == 8:  # 容差列
                cell.alignment = ALIGN_CENTER
            elif col in (9, 10):  # 公式和说明
                cell.alignment = ALIGN_LEFT

    # ── 科目追溯 sheet ──

    def _write_trace_sheet(
        self, ws: Worksheet, summary: ValidationSummary
    ) -> None:
        """写入科目追溯 sheet。"""
        headers = [
            "规则ID", "规则名称", "侧", "科目名称",
            "金额", "原始行", "原始列",
        ]
        self._write_header_row(ws, headers)

        row = 2
        for result in summary.results:
            if result.skipped:
                continue
            for trace_item in result.trace:
                side_label = "左" if trace_item.side == "left" else "右"
                values = [
                    result.rule_id,
                    result.rule_name,
                    side_label,
                    trace_item.name,
                    trace_item.amount,
                    _format_source_row(trace_item.row),
                    trace_item.column,
                ]
                for col, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.font = FONT_NORMAL
                    cell.border = THIN_BORDER
                    if col == 5:  # 金额列
                        cell.number_format = "#,##0.00"
                        cell.alignment = ALIGN_RIGHT
                    elif col in (6, 7):  # 行列（原始行可能为「第X页表内第N行」中文串）
                        cell.alignment = ALIGN_CENTER
                    else:
                        cell.alignment = ALIGN_LEFT
                row += 1

        # 列宽
        col_widths = [14, 28, 6, 24, 18, 14, 22]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = "A2"

    # ── 通用样式 ──

    @staticmethod
    def _write_header_row(ws: Worksheet, headers: list[str]) -> None:
        """写入表头行并应用样式。"""
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
