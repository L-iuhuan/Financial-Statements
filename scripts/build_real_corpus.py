"""真实年报语料库构建脚本 (路径 A: 东方财富 akshare 结构化数据)。

拉取 18 家 A 股上市公司年度报告的三大报表 (资产负债表/利润表/现金流量表),
转换为项目可导入的标准 Excel 格式, 存至 tests/fixtures/real_reports/, 并维护
manifest.json (公司/代码/行业/下载日期/数据年份/已知真实差异)。

合规红线:
- tests/fixtures/real_reports/ 已在 .gitignore 中忽略, 真实报表数据严禁提交 git。
- manifest.json 记录数据年份与已知真实差异, 供 validate_corpus.py 审查。

用法: python scripts/build_real_corpus.py [--year 2025]
"""

from __future__ import annotations

import argparse
import json
import math
import sys
from datetime import date
from numbers import Real
from pathlib import Path
from types import ModuleType

import openpyxl
import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet

sys.stdout.reconfigure(encoding="utf-8")  # 中文/✓✗ 输出 (GBK 控制台防乱码, 同 verify_sce.py 约定)

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
_OUTPUT_DIR = _PROJECT_ROOT / "tests" / "fixtures" / "real_reports"
_MANIFEST = _OUTPUT_DIR / "manifest.json"

# (akshare 代码, 股票代码, 公司名称, 行业)
COMPANIES: list[tuple[str, str, str, str]] = [
    ("SH601398", "601398", "工商银行", "financial"),
    ("SH600036", "600036", "招商银行", "financial"),
    ("SH601318", "601318", "中国平安", "financial"),
    ("SH600030", "600030", "中信证券", "financial"),
    ("SZ000002", "000002", "万科A", "real_estate"),
    ("SH601668", "601668", "中国建筑", "construction"),
    ("SH600519", "600519", "贵州茅台", "general"),
    ("SZ000858", "000858", "五粮液", "general"),
    ("SZ000651", "000651", "格力电器", "general"),
    ("SH601933", "601933", "永辉超市", "retail"),
    ("SH600276", "600276", "恒瑞医药", "high_growth"),
    ("SH600900", "600900", "长江电力", "general"),
    ("SH600019", "600019", "宝钢股份", "cyclical"),
    ("SH600309", "600309", "万华化学", "cyclical"),
    ("SZ002415", "002415", "海康威视", "high_growth"),
    ("SH600104", "600104", "上汽集团", "cyclical"),
    ("SZ300750", "300750", "宁德时代", "high_growth"),
    ("SH601012", "601012", "隆基绿能", "cyclical"),
]

# 东方财富英文字段 -> 项目可识别的中文科目名 (与 name_mapper 一致)。
# 字典插入顺序即报表行顺序。
BS_EM_TO_CN: dict[str, str] = {
    "MONETARYFUNDS": "货币资金",
    "TRADE_FINASSET": "交易性金融资产",
    "NOTE_RECE": "应收票据",
    "ACCOUNTS_RECE": "应收账款",
    "ADVANCE_RECEIVABLES": "预收款项",
    "PREPAYMENT": "预付款项",
    "TOTAL_OTHER_RECE": "其他应收款",
    "INVENTORY": "存货",
    "CONTRACT_ASSET": "合同资产",
    "TOTAL_CURRENT_ASSETS": "流动资产合计",
    "FIXED_ASSET": "固定资产",
    "CIP": "在建工程",
    "USERIGHT_ASSET": "使用权资产",
    "INTANGIBLE_ASSET": "无形资产",
    "GOODWILL": "商誉",
    "LONG_PREPAID_EXPENSE": "长期待摊费用",
    "DEFER_TAX_ASSET": "递延所得税资产",
    "TOTAL_NONCURRENT_ASSETS": "非流动资产合计",
    "TOTAL_ASSETS": "资产总计",
    "SHORT_LOAN": "短期借款",
    "NOTE_PAYABLE": "应付票据",
    "ACCOUNTS_PAYABLE": "应付账款",
    "CONTRACT_LIAB": "合同负债",
    "STAFF_SALARY_PAYABLE": "应付职工薪酬",
    "TAX_PAYABLE": "应交税费",
    "TOTAL_OTHER_PAYABLE": "其他应付款",
    "NONCURRENT_LIAB_1YEAR": "一年内到期的非流动负债",
    "TOTAL_CURRENT_LIAB": "流动负债合计",
    "LONG_LOAN": "长期借款",
    "BOND_PAYABLE": "应付债券",
    "LEASE_LIAB": "租赁负债",
    "DEFER_TAX_LIAB": "递延所得税负债",
    "TOTAL_NONCURRENT_LIAB": "非流动负债合计",
    "TOTAL_LIABILITIES": "负债合计",
    "SHARE_CAPITAL": "实收资本",
    "CAPITAL_RESERVE": "资本公积",
    "TREASURY_SHARES": "库存股",
    "OTHER_COMPRE_INCOME": "其他综合收益",
    "SURPLUS_RESERVE": "盈余公积",
    "UNASSIGN_RPOFIT": "未分配利润",
    "TOTAL_EQUITY": "所有者权益合计",
    "TOTAL_LIAB_EQUITY": "负债和所有者权益总计",
    "MINORITY_EQUITY": "少数股东权益",
    "TOTAL_PARENT_EQUITY": "归属于母公司所有者权益",
    "GENERAL_RISK_RESERVE": "一般风险准备",
    "SPECIAL_RESERVE": "专项储备",
    "OTHER_EQUITY_TOOL": "其他权益工具",
}

IS_EM_TO_CN: dict[str, str] = {
    "TOTAL_OPERATE_INCOME": "营业总收入",
    "OPERATE_INCOME": "营业收入",
    "INTEREST_INCOME": "利息收入",
    "EARNED_PREMIUM": "已赚保费",
    "FEE_COMMISSION_INCOME": "手续费及佣金收入",
    "OTHER_BUSINESS_INCOME": "其他业务收入",
    "TOTAL_OPERATE_COST": "营业总成本",
    "OPERATE_COST": "营业成本",
    "INTEREST_EXPENSE": "利息支出",
    "FEE_COMMISSION_EXPENSE": "手续费及佣金支出",
    "RESEARCH_EXPENSE": "研发费用",
    "SURRENDER_VALUE": "退保金",
    "NET_COMPENSATE_EXPENSE": "赔付支出",
    "NET_CONTRACT_RESERVE": "提取保险责任准备金净额",
    "POLICY_BONUS_EXPENSE": "保单红利支出",
    "REINSURE_EXPENSE": "分保费用",
    "OPERATE_TAX_ADD": "税金及附加",
    "SALE_EXPENSE": "销售费用",
    "MANAGE_EXPENSE": "管理费用",
    "FINANCE_EXPENSE": "财务费用",
    "FAIRVALUE_CHANGE_INCOME": "公允价值变动收益",
    "INVEST_INCOME": "投资收益",
    "ASSET_DISPOSAL_INCOME": "资产处置收益",
    "OTHER_INCOME": "其他收益",
    "OPERATE_PROFIT": "营业利润",
    "NONBUSINESS_INCOME": "营业外收入",
    "NONBUSINESS_EXPENSE": "营业外支出",
    "TOTAL_PROFIT": "利润总额",
    "INCOME_TAX": "所得税费用",
    "NETPROFIT": "净利润",
    "OTHER_COMPRE_INCOME": "税后其他综合收益",
    "TOTAL_COMPRE_INCOME": "综合收益总额",
    "DEDUCT_PARENT_NETPROFIT": "扣除非经常性损益的净利润",
}

# 利润表减值损失: (中文名, 收入型字段, 损失型字段)。
# 非金融企业走 *_INCOME (本身为负, 与报表一致); 金融企业走 *_LOSS (东财给正数, 报表为减项)。
IS_IMPAIRMENT_ROWS: list[tuple[str, str, str]] = [
    ("信用减值损失", "CREDIT_IMPAIRMENT_INCOME", "CREDIT_IMPAIRMENT_LOSS"),
    ("资产减值损失", "ASSET_IMPAIRMENT_INCOME", "ASSET_IMPAIRMENT_LOSS"),
]

CF_EM_TO_CN: dict[str, str] = {
    "SALES_SERVICES": "销售商品、提供劳务收到的现金",
    "RECEIVE_TAX_REFUND": "收到的税费返还",
    "RECEIVE_OTHER_OPERATE": "收到其他与经营活动有关的现金",
    "TOTAL_OPERATE_INFLOW": "经营活动现金流入小计",
    "BUY_SERVICES": "购买商品、接受劳务支付的现金",
    "PAY_STAFF_CASH": "支付给职工以及为职工支付的现金",
    "PAY_ALL_TAX": "支付的各项税费",
    "PAY_OTHER_OPERATE": "支付其他与经营活动有关的现金",
    "TOTAL_OPERATE_OUTFLOW": "经营活动现金流出小计",
    "NETCASH_OPERATE": "经营活动产生的现金流量净额",
    "WITHDRAW_INVEST": "收回投资收到的现金",
    "RECEIVE_INVEST_INCOME": "取得投资收益收到的现金",
    "DISPOSAL_LONG_ASSET": "处置固定资产、无形资产和其他长期资产收回的现金净额",
    "TOTAL_INVEST_INFLOW": "投资活动现金流入小计",
    "CONSTRUCT_LONG_ASSET": "购建固定资产、无形资产和其他长期资产支付的现金",
    "INVEST_PAY_CASH": "投资支付的现金",
    "TOTAL_INVEST_OUTFLOW": "投资活动现金流出小计",
    "NETCASH_INVEST": "投资活动产生的现金流量净额",
    "RECEIVE_LOAN_CASH": "取得借款收到的现金",
    "TOTAL_FINANCE_INFLOW": "筹资活动现金流入小计",
    "PAY_DEBT_CASH": "偿还债务支付的现金",
    "ASSIGN_DIVIDEND_PORFIT": "分配股利、利润或偿付利息支付的现金",
    "TOTAL_FINANCE_OUTFLOW": "筹资活动现金流出小计",
    "NETCASH_FINANCE": "筹资活动产生的现金流量净额",
    "RATE_CHANGE_EFFECT": "汇率变动对现金及现金等价物的影响",
    "CCE_ADD": "现金及现金等价物净增加额",
    "BEGIN_CCE": "期初现金及现金等价物余额",
    "END_CCE": "期末现金及现金等价物余额",
}

# 现金流量表"补充资料"区域 (映射为 cf_notes_ 前缀变量)
CF_SUPPLEMENTARY_EM_TO_CN: dict[str, str] = {
    "NETPROFIT": "净利润",
    "ASSET_IMPAIRMENT": "资产减值准备",
    "FA_IR_DEPR": "固定资产折旧",
    "IA_AMORTIZE": "无形资产摊销",
    "LPE_AMORTIZE": "长期待摊费用摊销",
    "DISPOSAL_LONGASSET_LOSS": "处置固定资产、无形资产和其他长期资产的损失",
    "FA_SCRAP_LOSS": "固定资产报废损失",
    "FAIRVALUE_CHANGE_LOSS": "公允价值变动损失",
    "FINANCE_EXPENSE": "财务费用",
    "INVEST_LOSS": "投资损失",
    "DT_ASSET_REDUCE": "递延所得税资产减少",
    "DT_LIAB_ADD": "递延所得税负债增加",
    "INVENTORY_REDUCE": "存货的减少",
    "OPERATE_RECE_REDUCE": "经营性应收项目的减少",
    "OPERATE_PAYABLE_ADD": "经营性应付项目的增加",
    "NETCASH_OPERATENOTE": "经营活动产生的现金流量净额",
    "CCE_ADDNOTE": "现金及现金等价物净增加额",
    "END_CASH": "现金的期末余额",
    "BEGIN_CASH": "现金的期初余额",
}

_SUPPLEMENTARY_MARKER = "补充资料"


def _to_number(value: object) -> float | None:
    """将单元格值转换为数字, 空值/NaN/不可解析返回 None。

    兼容 numpy 标量 (int64/float64): numpy 数值均注册为 numbers.Real。
    """
    if value is None or isinstance(value, bool):
        return None
    if not isinstance(value, Real):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if math.isnan(number):
        return None
    return number


def _impairment_value(row: dict[str, object], income_field: str, loss_field: str) -> float | None:
    """读取减值损失金额, 统一为负值 (与报表列示方向一致)。

    非金融企业用 *_INCOME 字段 (本身为负); 金融企业用 *_LOSS 字段
    (东财给正数绝对值, 报表列示为减项, 故取负)。
    """
    income = _to_number(row.get(income_field))
    if income is not None:
        return income
    loss = _to_number(row.get(loss_field))
    if loss is not None:
        return -loss
    return None


def _fetch_statement(ak_module: ModuleType, func_name: str, symbol: str) -> pd.DataFrame:
    """调用 akshare 函数获取某张报表全部报告期数据。"""
    func = getattr(ak_module, func_name)
    result = func(symbol=symbol)
    if not isinstance(result, pd.DataFrame):
        raise TypeError(f"akshare 函数 {func_name} 未返回 DataFrame")
    return result


def _extract_year_rows(
    df: pd.DataFrame, year: int, report_type: str = "年报"
) -> tuple[dict[str, object] | None, dict[str, object] | None]:
    """从全报告期数据中提取指定年份年报行 (本年) 与上一年年报行 (上年)。

    Args:
        df: akshare 返回的 DataFrame
        year: 目标数据年份
        report_type: 报表类型, 默认取"年报"

    Returns:
        (本年行, 上年行) 字典; 缺失年份返回 None
    """
    current_mask = df["REPORT_DATE"].astype(str).str.startswith(f"{year}-12-31")
    prior_mask = df["REPORT_DATE"].astype(str).str.startswith(f"{year - 1}-12-31")
    df_current: pd.DataFrame = df.loc[current_mask]
    df_prior: pd.DataFrame = df.loc[prior_mask]
    if report_type:
        df_current = df_current.loc[df_current["REPORT_TYPE"] == report_type]
        df_prior = df_prior.loc[df_prior["REPORT_TYPE"] == report_type]
    current: dict[str, object] | None = None
    prior: dict[str, object] | None = None
    if not df_current.empty:
        current = dict(df_current.iloc[0])
    if not df_prior.empty:
        prior = dict(df_prior.iloc[0])
    return current, prior


def _build_row_data(
    mapping: dict[str, str], current: dict[str, object], prior: dict[str, object]
) -> list[tuple[str, float | None, float | None]]:
    """按映射生成 [(中文科目名, 本年值, 上年值)] 列表, 跳过空值行。"""
    rows: list[tuple[str, float | None, float | None]] = []
    for em_field, cn_name in mapping.items():
        cur_value = _to_number(current.get(em_field))
        pri_value = _to_number(prior.get(em_field))
        if cur_value is None and pri_value is None:
            continue
        rows.append((cn_name, cur_value, pri_value))
    return rows


def _build_is_row_data(
    current: dict[str, object], prior: dict[str, object]
) -> list[tuple[str, float | None, float | None]]:
    """构建利润表行数据, 含减值损失的字段选择与符号归一。"""
    rows: list[tuple[str, float | None, float | None]] = []
    for em_field, cn_name in IS_EM_TO_CN.items():
        cur_value = _to_number(current.get(em_field))
        pri_value = _to_number(prior.get(em_field))
        if cur_value is None and pri_value is None:
            continue
        rows.append((cn_name, cur_value, pri_value))
    for cn_name, income_field, loss_field in IS_IMPAIRMENT_ROWS:
        cur_value = _impairment_value(current, income_field, loss_field)
        pri_value = _impairment_value(prior, income_field, loss_field)
        if cur_value is None and pri_value is None:
            continue
        rows.append((cn_name, cur_value, pri_value))
    return rows


def _active_sheet(wb: openpyxl.Workbook) -> Worksheet:
    """获取工作簿的活动工作表。"""
    sheet = wb.active
    assert sheet is not None
    return sheet


def _write_statement_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    title: str,
    table_no: str,
    header: list[str],
    rows: list[tuple[str, float | None, float | None]],
    company_name: str,
    year: int,
) -> None:
    """写入一张报表工作表 (与现有测试报表 fixture 格式一致)。"""
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
    ws.cell(row=1, column=1, value=f"编制单位：{company_name}")
    ws.cell(row=2, column=1, value=title)
    ws.cell(row=2, column=3, value=table_no)
    period_text = f"{year}年12月31日" if "资产" in title else f"{year}年度"
    ws.cell(row=3, column=1, value=period_text)
    ws.cell(row=3, column=3, value="单位：元")
    for col_idx, header_text in enumerate(header, 1):
        ws.cell(row=4, column=col_idx, value=header_text)
    for row_idx, (name, cur, pri) in enumerate(rows, 5):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=row_idx - 4)
        ws.cell(row=row_idx, column=3, value=cur)
        ws.cell(row=row_idx, column=4, value=pri)


def _write_cf_sheet(
    wb: openpyxl.Workbook,
    main_rows: list[tuple[str, float | None, float | None]],
    supp_rows: list[tuple[str, float | None, float | None]],
    company_name: str,
    year: int,
) -> None:
    """写入现金流量表工作表, 含"补充资料"区域。"""
    ws = wb.create_sheet("现金流量表")
    ws.cell(row=1, column=1, value=f"编制单位：{company_name}")
    ws.cell(row=2, column=1, value="现金流量表")
    ws.cell(row=2, column=3, value="会企03表")
    ws.cell(row=3, column=1, value=f"{year}年度")
    ws.cell(row=3, column=3, value="单位：元")
    header = ["项目", "行次", "本期金额", "上期金额"]
    for col_idx, header_text in enumerate(header, 1):
        ws.cell(row=4, column=col_idx, value=header_text)
    all_rows = list(main_rows)
    if supp_rows:
        all_rows.append((_SUPPLEMENTARY_MARKER, None, None))
        all_rows.extend(supp_rows)
    for row_idx, (name, cur, pri) in enumerate(all_rows, 5):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=row_idx - 4)
        ws.cell(row=row_idx, column=3, value=cur)
        ws.cell(row=row_idx, column=4, value=pri)


def _build_company_excel(
    ak_module: ModuleType,
    symbol: str,
    company_name: str,
    year: int,
) -> tuple[Path | None, str]:
    """为一家公司生成三大报表 Excel 文件。

    Returns:
        (输出文件路径, 说明); 失败时路径为 None
    """
    try:
        bs_df = _fetch_statement(ak_module, "stock_balance_sheet_by_report_em", symbol)
        is_df = _fetch_statement(ak_module, "stock_profit_sheet_by_report_em", symbol)
        cf_df = _fetch_statement(ak_module, "stock_cash_flow_sheet_by_report_em", symbol)
    except Exception as exc:
        return None, f"拉取报表数据失败: {type(exc).__name__}: {exc}"

    bs_current, bs_prior = _extract_year_rows(bs_df, year)
    is_current, is_prior = _extract_year_rows(is_df, year)
    cf_current, cf_prior = _extract_year_rows(cf_df, year)

    if bs_current is None or bs_prior is None:
        return None, f"资产负债表缺少 {year} 或 {year - 1} 年报数据"
    if is_current is None or is_prior is None:
        return None, f"利润表缺少 {year} 或 {year - 1} 年报数据"
    if cf_current is None or cf_prior is None:
        return None, f"现金流量表缺少 {year} 或 {year - 1} 年报数据"

    wb = openpyxl.Workbook()
    _active_sheet(wb).title = "资产负债表"
    _write_statement_sheet(
        wb, "资产负债表", "资产负债表", "会企01表",
        ["项目", "行次", "期末数", "年初数"],
        _build_row_data(BS_EM_TO_CN, bs_current, bs_prior),
        company_name, year,
    )
    _write_statement_sheet(
        wb, "利润表", "利润表", "会企02表",
        ["项目", "行次", "本期数", "上期数"],
        _build_is_row_data(is_current, is_prior),
        company_name, year,
    )
    _write_cf_sheet(
        wb,
        _build_row_data(CF_EM_TO_CN, cf_current, cf_prior),
        _build_row_data(CF_SUPPLEMENTARY_EM_TO_CN, cf_current, cf_prior),
        company_name, year,
    )
    output = _OUTPUT_DIR / f"{company_name}_{year}年报_三大报表.xlsx"
    wb.save(str(output))
    return output, f"资产{bs_current.get('TOTAL_ASSETS')} 负债{bs_current.get('TOTAL_LIABILITIES')}"


def _load_manifest() -> dict:
    """读取已有 manifest, 不存在返回空结构。"""
    if _MANIFEST.exists():
        return json.loads(_MANIFEST.read_text(encoding="utf-8"))
    return {"schema_version": 1, "companies": {}}


def _update_manifest(
    manifest: dict,
    code: str,
    name: str,
    industry: str,
    year: int,
    status: str,
    note: str,
) -> None:
    """更新 manifest 中的一家公司记录。"""
    companies = manifest.setdefault("companies", {})
    record = companies.setdefault(code, {})
    record.update(
        {
            "code": code,
            "name": name,
            "industry": industry,
            "download_date": date.today().isoformat(),
            "data_year": year,
            "status": status,
            "note": note,
            "known_real_diffs": record.get("known_real_diffs", []),
        }
    )


def main() -> None:
    """构建语料库主流程。"""
    parser = argparse.ArgumentParser(description="构建真实年报语料库 (akshare 路径)")
    parser.add_argument("--year", type=int, default=2025, help="目标数据年份, 默认 2025")
    parser.add_argument("--code", type=str, default="", help="仅构建指定公司代码")
    args = parser.parse_args()

    _OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    manifest = _load_manifest()

    import akshare as ak

    for symbol, code, name, industry in COMPANIES:
        if args.code and code != args.code:
            continue
        output, note = _build_company_excel(ak, symbol, name, args.year)
        if output is None:
            print(f"  ✗ {name}({code}): {note}")
            _update_manifest(manifest, code, name, industry, args.year, "failed", note)
            continue
        _update_manifest(manifest, code, name, industry, args.year, "ok", note)
        print(f"  ✓ {name}({code}): {output.name} ({note})")

    _MANIFEST.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"\nmanifest 已更新: {_MANIFEST}")


if __name__ == "__main__":
    main()

