"""审计底稿导出验证脚本。

用法: python scripts/verify_export.py

流程:
1. 对贵州茅台 2023 年报执行校验
2. 导出为 Excel 审计底稿
3. 回读验证 sheet 存在性和行数
4. 打印 PASS/FAIL
"""

from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from fsa.core.engine.registry import RuleRegistry
from fsa.core.exporter.audit_exporter import AuditExporter
from fsa.core.importer.importer import ImportService
from fsa.services.validation_service import ValidationService


def main() -> None:
    base = Path(__file__).resolve().parent.parent
    report_path = base / "tests/fixtures/real_reports/贵州茅台_2023年报_三大报表.xlsx"

    if not report_path.exists():
        print(f"FAIL: 报表文件不存在: {report_path}")
        sys.exit(1)

    # 1. 导入报表
    print("1. 导入报表...")
    importer = ImportService(period="2023-12")
    reports = importer.import_file(str(report_path))
    print(f"   导入 {len(reports)} 张报表")

    # 2. 执行校验
    print("2. 执行校验...")
    registry = RuleRegistry.from_json("cas_gouji_rule_library.json")
    service = ValidationService(registry)
    summary = service.validate(reports, "2023-12")
    print(f"   总 {summary.total} 条, 通过 {summary.passed}, 不通过 {summary.failed}, 异常 {summary.errored}, 跳过 {summary.skipped}")

    # 3. 导出
    print("3. 导出审计底稿...")
    exporter = AuditExporter()
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp_path = f.name
    try:
        exporter.export(summary, tmp_path)
        print(f"   导出到: {tmp_path}")

        # 4. 验证
        print("4. 验证导出文件...")
        wb = load_workbook(tmp_path)

        errors: list[str] = []

        # 检查 sheet 存在
        expected_sheets = ["校验汇总", "校验明细", "科目追溯"]
        for sheet_name in expected_sheets:
            if sheet_name not in wb.sheetnames:
                errors.append(f"缺少 sheet: {sheet_name}")

        # 检查明细 sheet 行数
        ws_detail = wb["校验明细"]
        detail_rows = ws_detail.max_row - 1
        expected_detail = len(summary.results)
        if detail_rows != expected_detail:
            errors.append(f"明细行数不匹配: 预期 {expected_detail}, 实际 {detail_rows}")

        # 检查追溯 sheet 行数
        ws_trace = wb["科目追溯"]
        trace_rows = ws_trace.max_row - 1
        expected_trace = sum(
            len(r.trace) for r in summary.results if not r.skipped
        )
        if trace_rows != expected_trace:
            errors.append(f"追溯行数不匹配: 预期 {expected_trace}, 实际 {trace_rows}")

        # 检查汇总 sheet
        ws_summary = wb["校验汇总"]
        summary_values = " ".join(
            str(cell) for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row, values_only=True)
            for cell in row if cell is not None
        )
        if "2023-12" not in summary_values:
            errors.append("汇总 sheet 缺少报告期间")

        if "1.1.1" not in summary_values:
            errors.append("汇总 sheet 缺少规则库版本")

        wb.close()

        if errors:
            print(f"\nFAIL: {len(errors)} 个问题:")
            for e in errors:
                print(f"  - {e}")
            sys.exit(1)
        else:
            print("\nPASS: 所有验证通过")
            print("  - 3 个 sheet 均存在")
            print(f"  - 明细行数: {detail_rows}")
            print(f"  - 追溯行数: {trace_rows}")
            print("  - 汇总信息完整")
    finally:
        os.unlink(tmp_path)


if __name__ == "__main__":
    main()
