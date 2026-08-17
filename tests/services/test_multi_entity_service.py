"""多主体批量校验与双边核对测试。"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from fsa.core.engine.registry import RuleRegistry
from fsa.core.exceptions import FSAError
from fsa.core.importer.detail_importer import DetailImporter
from fsa.core.importer.importer import ImportService
from fsa.core.models.detail import DetailDataset
from fsa.core.models.report import Report, ReportType
from fsa.core.models.result import ValidationResult
from fsa.services.entity_config import EntityConfig, load_entity_configs
from fsa.services.multi_entity_service import MultiEntityService


def _write_detail(path: Path, entity: str, counterparty: str, project: str, amount: float) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "内部现金流量明细表"
    headers = ["月份", "统计单位名称", "对方单位名称", "款项性质", "现金流量项目", "发生额"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    row = [1, entity, counterparty, "货款", project, amount]
    for col_idx, value in enumerate(row, 1):
        ws.cell(row=2, column=col_idx, value=value)
    wb.save(str(path))


def _make_entity_folder(
    tmp_path: Path, name: str, counterparty: str, project: str, amount: float
) -> Path:
    folder = tmp_path / name
    folder.mkdir()
    from tests.importer.conftest import make_multi_sheet_excel

    make_multi_sheet_excel(folder)
    _write_detail(folder / "detail.xlsx", name, counterparty, project, amount)
    return folder


def _write_dar_balance_sheet(
    folder: Path, asset: float = 100.0, liability: float = 88.0
) -> Path:
    """写入指定资产负债率 (liability/asset) 的资产负债表 Excel。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "资产负债表"
    headers = ["项目", "行次", "期末余额", "年初余额"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    rows = [
        ("资产总计", 20, asset, asset),
        ("负债合计", 35, liability, liability),
        ("所有者权益合计", 50, asset - liability, asset - liability),
    ]
    for row_idx, (name, row_num, ending, beginning) in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=row_num)
        ws.cell(row=row_idx, column=3, value=ending)
        ws.cell(row=row_idx, column=4, value=beginning)
    path = folder / "balance_sheet.xlsx"
    wb.save(str(path))
    return path


def _make_dar_folder(
    tmp_path: Path, name: str, asset: float = 100.0, liability: float = 88.0
) -> Path:
    """创建只含资产负债率 0.88 资产负债表的实体文件夹。"""
    folder = tmp_path / name
    folder.mkdir()
    _write_dar_balance_sheet(folder, asset, liability)
    return folder


def _registry() -> RuleRegistry:
    return RuleRegistry.from_json(
        Path(__file__).resolve().parent.parent.parent / "cas_gouji_rule_library.json"
    )


class TestMultiEntityService:
    """多主体批量校验。"""

    def test_validate_two_entities(self, tmp_path: Path) -> None:
        folder_a = _make_entity_folder(
            tmp_path, "杭州杰为", "拓尔微", "收到的其他与经营活动的现金", 100.0
        )
        folder_b = _make_entity_folder(
            tmp_path, "拓尔微", "杭州杰为", "支付的其他与经营活动的现金", 100.0
        )
        result = MultiEntityService(_registry()).validate_folders(
            [str(folder_a), str(folder_b)], period="2026-06"
        )
        assert len(result.outcomes) == 2
        assert result.combined is not None
        assert result.combined.total > 0
        assert len(result.bilateral) == 1
        assert result.bilateral[0].passed is True

    def test_bilateral_mismatch_reported(self, tmp_path: Path) -> None:
        folder_a = _make_entity_folder(
            tmp_path, "杭州杰为", "拓尔微", "收到的其他与经营活动的现金", 120.0
        )
        folder_b = _make_entity_folder(
            tmp_path, "拓尔微", "杭州杰为", "支付的其他与经营活动的现金", 100.0
        )
        result = MultiEntityService(_registry()).validate_folders(
            [str(folder_a), str(folder_b)], period="2026-06"
        )
        bilateral = next(
            r for r in result.bilateral if r.rule_id == "ICF-002"
        )
        assert bilateral.passed is False
        assert bilateral.diff == 20.0


class TestBilateralExtension:
    """内部现金流双边核对: 科目对扩展与配置覆写。"""

    def test_investing_pair_matches(self, tmp_path: Path) -> None:
        """投资活动其他收支科目对纳入核对。"""
        folder_a = _make_entity_folder(
            tmp_path, "杭州杰为", "拓尔微", "收到的其他与投资活动的现金", 100.0
        )
        folder_b = _make_entity_folder(
            tmp_path, "拓尔微", "杭州杰为", "支付的其他与投资活动的现金", 100.0
        )
        result = MultiEntityService(_registry()).validate_folders(
            [str(folder_a), str(folder_b)], period="2026-06"
        )
        assert len(result.bilateral) == 1
        assert result.bilateral[0].passed is True
        assert "投资活动" in result.bilateral[0].message

    def test_financing_pair_matches(self, tmp_path: Path) -> None:
        """筹资活动其他收支科目对纳入核对。"""
        folder_a = _make_entity_folder(
            tmp_path, "杭州杰为", "拓尔微", "收到的其他与筹资活动的现金", 200.0
        )
        folder_b = _make_entity_folder(
            tmp_path, "拓尔微", "杭州杰为", "支付的其他与筹资活动的现金", 200.0
        )
        result = MultiEntityService(_registry()).validate_folders(
            [str(folder_a), str(folder_b)], period="2026-06"
        )
        assert len(result.bilateral) == 1
        assert result.bilateral[0].passed is True
        assert "筹资活动" in result.bilateral[0].message

    def test_bilateral_tolerance_override(self, tmp_path: Path) -> None:
        """entity_config.bilateral_tolerance 覆写: 差额 2.0 在容差 5.0 内通过。"""
        folder_a = _make_entity_folder(
            tmp_path, "杭州杰为", "拓尔微", "收到的其他与经营活动的现金", 100.0
        )
        folder_b = _make_entity_folder(
            tmp_path, "拓尔微", "杭州杰为", "支付的其他与经营活动的现金", 98.0
        )
        configs = {
            "杭州杰为": EntityConfig(entity_id="杭州杰为", bilateral_tolerance=5.0),
        }
        result = MultiEntityService(_registry(), configs=configs).validate_folders(
            [str(folder_a), str(folder_b)], period="2026-06"
        )
        assert len(result.bilateral) == 1
        assert result.bilateral[0].passed is True
        assert result.bilateral[0].tolerance == 5.0

    def test_bilateral_pairs_override_effective(self, tmp_path: Path) -> None:
        """entity_config.bilateral_pairs 覆写: 自定义科目对参与核对。"""
        folder_a = _make_entity_folder(
            tmp_path, "甲", "乙", "收到其他与投资活动有关的现金", 100.0
        )
        folder_b = _make_entity_folder(
            tmp_path, "乙", "甲", "支付其他与投资活动有关的现金", 100.0
        )
        configs = {
            "甲": EntityConfig(
                entity_id="甲",
                bilateral_pairs={
                    "收到其他与投资活动有关的现金": "支付其他与投资活动有关的现金"
                },
            )
        }
        result = MultiEntityService(_registry(), configs=configs).validate_folders(
            [str(folder_a), str(folder_b)], period="2026-06"
        )
        assert len(result.bilateral) == 1
        assert result.bilateral[0].passed is True


class TestEntityConfig:
    """主体级口径配置。"""

    def test_load_and_convert_config(self, tmp_path: Path) -> None:
        path = tmp_path / "entity_configs.json"
        path.write_text(
            '{"entities": {"杭州杰为": {"tolerance": 0.05, '
            '"cash_equivalent_codes": ["1002"], '
            '"tb_to_bs_mappings": {"monetary_funds": '
            '{"codes": ["1002"], "side": "debit"}}}}}',
            encoding="utf-8",
        )
        configs = load_entity_configs(path)
        config = configs["杭州杰为"]
        assert config.tolerance == 0.05
        assert config.cash_equivalent_codes == ("1002",)
        detail_config = config.to_detail_config()
        assert detail_config.tolerance == 0.05
        assert detail_config.cash_equivalent_codes == ("1002",)

    def test_default_config_keeps_standard_mappings(self) -> None:
        config = EntityConfig(entity_id="主体A")
        detail_config = config.to_detail_config()
        assert "monetary_funds" in detail_config.tb_to_bs_mappings

    def test_industry_default_general(self) -> None:
        """industry 默认 general, 阈值与 general 一致。"""
        config = EntityConfig(entity_id="主体A")
        assert config.industry == "general"
        assert config.threshold_vars()["dar_threshold"] == 0.85

    def test_industry_loaded_and_applied(self, tmp_path: Path) -> None:
        """industry 从配置 JSON 加载并映射阈值。"""
        path = tmp_path / "entity_configs.json"
        path.write_text(
            '{"entities": {"银行": {"industry": "financial"}}}',
            encoding="utf-8",
        )
        configs = load_entity_configs(path)
        config = configs["银行"]
        assert config.industry == "financial"
        assert config.threshold_vars()["dar_threshold"] == 0.92

    def test_industry_threshold_vars_mapping(self) -> None:
        """各行业阈值映射抽查。"""
        from fsa.core.engine.thresholds import threshold_vars_for

        assert threshold_vars_for("construction")["ar_to_revenue_threshold"] == 0.60
        assert threshold_vars_for("construction")["sales_cash_ratio_threshold"] == 0.5
        assert threshold_vars_for("retail")["current_ratio_threshold"] == 0.7
        assert threshold_vars_for("cyclical")["gm_yoy_threshold"] == 0.50
        assert threshold_vars_for("high_growth")["yoy_fluctuation_threshold"] == 0.50

    def test_load_entity_configs_parses_new_fields(self, tmp_path: Path) -> None:
        """load_entity_configs 解析新增的科目对/容差字段并透传到 detail 配置。"""
        path = tmp_path / "entity_configs.json"
        path.write_text(
            '{"entities": {"主体A": {'
            '"reclass_pairs": {"应收账款": ["预付款项"]}, '
            '"balance_sheet_accounts": {"accounts_receivable": "应收账款"}, '
            '"margin_tolerance": 0.02, '
            '"bilateral_pairs": {"收到的其他与投资活动的现金": "支付的其他与投资活动的现金"}, '
            '"bilateral_tolerance": 0.05}}}',
            encoding="utf-8",
        )
        configs = load_entity_configs(path)
        config = configs["主体A"]
        assert config.reclass_pairs == {"应收账款": ("预付款项",)}
        assert config.balance_sheet_accounts == {"accounts_receivable": "应收账款"}
        assert config.margin_tolerance == 0.02
        assert config.bilateral_pairs == {
            "收到的其他与投资活动的现金": "支付的其他与投资活动的现金"
        }
        assert config.bilateral_tolerance == 0.05
        detail_config = config.to_detail_config()
        assert detail_config.reclass_pairs == {"应收账款": ("预付款项",)}
        assert detail_config.balance_sheet_accounts == {
            "accounts_receivable": "应收账款"
        }
        assert detail_config.margin_tolerance == 0.02

    def test_new_fields_default_to_none(self) -> None:
        """新增字段缺省为 None（不改变默认行为）。"""
        config = EntityConfig(entity_id="主体A")
        assert config.reclass_pairs is None
        assert config.balance_sheet_accounts is None
        assert config.margin_tolerance is None
        assert config.bilateral_pairs is None
        assert config.bilateral_tolerance is None
        detail_config = config.to_detail_config()
        assert detail_config.reclass_pairs is None
        assert detail_config.balance_sheet_accounts is None
        assert detail_config.margin_tolerance is None


class TestMultiEntityIndustryThresholds:
    """多主体按行业注入 LR-* 阈值 (实体配置 -> 校验链路穿线)。"""

    @staticmethod
    def _dar_result(summary) -> ValidationResult:
        """从汇总结果中取 LR-DAR-001 结果。"""
        assert summary is not None
        return next(r for r in summary.results if r.rule_id == "LR-DAR-001")

    def test_financial_passes_general_fails(self, tmp_path: Path) -> None:
        """同一份资产负债率 0.88 报表: financial 主体通过, general 主体不通过。"""
        folder_bank = _make_dar_folder(tmp_path, "银行主体")
        folder_company = _make_dar_folder(tmp_path, "一般主体")
        configs = {
            "银行主体": EntityConfig(entity_id="银行主体", industry="financial"),
        }
        result = MultiEntityService(_registry(), configs=configs).validate_folders(
            [str(folder_bank), str(folder_company)], period="2026-06"
        )
        outcomes = {outcome.entity_id: outcome for outcome in result.outcomes}
        assert "银行主体" in outcomes
        assert "一般主体" in outcomes

        bank_dar = self._dar_result(outcomes["银行主体"].summary)
        company_dar = self._dar_result(outcomes["一般主体"].summary)
        assert bank_dar.passed is True   # financial 0.92 >= 0.88
        assert company_dar.passed is False  # general 0.85 < 0.88

    def test_single_entity_default_general_regression(self, tmp_path: Path) -> None:
        """单体校验 (无 entity_config) 走 general 默认阈值: 0.88 不通过。"""
        folder = _make_dar_folder(tmp_path, "默认主体")
        outcome = MultiEntityService(_registry()).validate_folder(
            str(folder), period="2026-06"
        )
        assert outcome.summary is not None
        dar = self._dar_result(outcome.summary)
        assert dar.passed is False


class TestImportOne:
    """_import_one 双轨导入: 主表成功后才尝试明细, 明细失败仅调试日志。"""

    def test_main_failure_records_single_error_and_skips_detail(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """主表导入失败: 记录一条错误且不再尝试明细导入（避免重复错误）。"""
        folder = tmp_path / "主体A"
        folder.mkdir()
        (folder / "bad.xlsx").write_text("非报表文件", encoding="utf-8")

        def _fail_main(self, file_path: str) -> list[Report]:
            raise FSAError(f"无法识别报表: {file_path}")

        detail_calls: list[str] = []

        def _spy_detail(self, file_path: str) -> DetailDataset:
            detail_calls.append(file_path)
            return DetailDataset()

        monkeypatch.setattr(ImportService, "import_file", _fail_main)
        monkeypatch.setattr(DetailImporter, "import_file", _spy_detail)

        outcome = MultiEntityService(_registry()).validate_folder(
            str(folder), period="2026-06"
        )

        assert len(outcome.errors) == 1
        assert "bad.xlsx" in outcome.errors[0]
        assert detail_calls == []

    def test_detail_failure_only_logged_debug(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """主表导入成功、明细导入失败: 不记录错误（纯主表文件属预期路径）。"""
        folder = tmp_path / "主体B"
        folder.mkdir()
        (folder / "main.xlsx").write_text("主表文件占位", encoding="utf-8")

        def _ok_main(self, file_path: str) -> list[Report]:
            return []

        def _fail_detail(self, file_path: str) -> DetailDataset:
            raise FSAError(f"非明细文件: {file_path}")

        monkeypatch.setattr(ImportService, "import_file", _ok_main)
        monkeypatch.setattr(DetailImporter, "import_file", _fail_detail)

        outcome = MultiEntityService(_registry()).validate_folder(
            str(folder), period="2026-06"
        )

        assert outcome.errors == []

    def test_main_success_then_detail_import(
        self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
    ) -> None:
        """主表导入成功后继续明细导入, 数据均被收集。"""
        folder = tmp_path / "主体C"
        folder.mkdir()
        (folder / "data.xlsx").write_text("占位", encoding="utf-8")

        captured: dict[str, list[str]] = {}

        def _ok_main(self, file_path: str) -> list[Report]:
            report = Report(
                report_type=ReportType.BALANCE_SHEET,
                period="2026-06",
                source_file=file_path,
                items=[],
            )
            return [report]

        def _spy_detail(self, file_path: str) -> DetailDataset:
            captured["detail"] = [file_path]
            return DetailDataset()

        monkeypatch.setattr(ImportService, "import_file", _ok_main)
        monkeypatch.setattr(DetailImporter, "import_file", _spy_detail)

        outcome = MultiEntityService(_registry()).validate_folder(
            str(folder), period="2026-06"
        )

        assert outcome.errors == []
        assert len(outcome.reports) == 1
        assert outcome.reports[0].report_type == ReportType.BALANCE_SHEET
        assert captured["detail"] == [str(folder / "data.xlsx")]
