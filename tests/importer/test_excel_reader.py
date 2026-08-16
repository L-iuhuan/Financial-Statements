"""excel_reader 模块测试: 读取 Excel 文件返回原始数据。

测试内容: 正常读取、文件不存在、空文件、合并单元格、多 sheet。
"""

from __future__ import annotations

from pathlib import Path

import pytest

from fsa.core.exceptions import FSAError
from fsa.core.importer.excel_reader import read_excel


class TestReadExcelNormal:
    """测试正常读取 Excel 文件。"""

    def test_read_balance_sheet_returns_correct_sheet_name(self) -> None:
        from tests.importer.conftest import make_balance_sheet_excel

        path = make_balance_sheet_excel()
        data = read_excel(str(path))

        assert "资产负债表" in data

    def test_read_balance_sheet_has_headers(self) -> None:
        from tests.importer.conftest import make_balance_sheet_excel

        path = make_balance_sheet_excel()
        data = read_excel(str(path))

        sheet = data["资产负债表"]
        assert "项目" in sheet.headers
        assert "期末余额" in sheet.headers
        assert "年初余额" in sheet.headers

    def test_read_balance_sheet_has_rows(self) -> None:
        from tests.importer.conftest import make_balance_sheet_excel

        path = make_balance_sheet_excel()
        data = read_excel(str(path))

        sheet = data["资产负债表"]
        assert len(sheet.rows) > 0

    def test_read_balance_sheet_row_has_correct_data(self) -> None:
        from tests.importer.conftest import make_balance_sheet_excel

        path = make_balance_sheet_excel()
        data = read_excel(str(path))

        sheet = data["资产负债表"]
        row = sheet.rows[0]
        assert row["项目"] == "流动资产合计"
        assert row["期末余额"] == 500000.00
        assert row["_row"] == 2

    def test_read_balance_sheet_asset_total_correct(self) -> None:
        from tests.importer.conftest import make_balance_sheet_excel

        path = make_balance_sheet_excel()
        data = read_excel(str(path))

        sheet = data["资产负债表"]
        asset_row = next(r for r in sheet.rows if r["项目"] == "资产总计")
        assert asset_row["期末余额"] == 2000000.00
        assert asset_row["年初余额"] == 1850000.00

    def test_read_income_statement_returns_correct_headers(self) -> None:
        from tests.importer.conftest import make_income_statement_excel

        path = make_income_statement_excel()
        data = read_excel(str(path))

        sheet = data["利润表"]
        assert "本期金额" in sheet.headers
        assert "上期金额" in sheet.headers

    def test_read_income_statement_net_profit_correct(self) -> None:
        from tests.importer.conftest import make_income_statement_excel

        path = make_income_statement_excel()
        data = read_excel(str(path))

        sheet = data["利润表"]
        np_row = next(r for r in sheet.rows if r["项目"] == "净利润")
        assert np_row["本期金额"] == 405000.00

    def test_read_cash_flow_statement_correct(self) -> None:
        from tests.importer.conftest import make_cash_flow_excel

        path = make_cash_flow_excel()
        data = read_excel(str(path))

        sheet = data["现金流量表"]
        operating_row = next(r for r in sheet.rows if r["项目"] == "经营活动产生的现金流量净额")
        assert operating_row["本期金额"] == 500000.00

    def test_read_multi_sheet_returns_all_sheets(self) -> None:
        from tests.importer.conftest import make_multi_sheet_excel

        path = make_multi_sheet_excel()
        data = read_excel(str(path))

        assert "资产负债表" in data
        assert "利润表" in data
        assert "现金流量表" in data
        assert len(data) == 3

    def test_read_multi_sheet_each_has_data(self) -> None:
        from tests.importer.conftest import make_multi_sheet_excel

        path = make_multi_sheet_excel()
        data = read_excel(str(path))

        for sheet in data.values():
            assert len(sheet.rows) > 0
            assert len(sheet.headers) > 0


class TestReadExcelErrors:
    """测试错误处理。"""

    def test_read_nonexistent_file_raises_error(self) -> None:
        with pytest.raises(FileNotFoundError):
            read_excel("nonexistent_file.xlsx")

    def test_read_empty_excel_returns_empty_rows(self) -> None:
        from tests.importer.conftest import make_empty_excel

        path = make_empty_excel()
        data = read_excel(str(path))

        assert len(data) == 1
        sheet = list(data.values())[0]
        assert len(sheet.rows) == 0


class TestReadExcelMergedCells:
    """测试合并单元格处理。"""

    def test_read_merged_cells_still_returns_data(self) -> None:
        from tests.importer.conftest import make_excel_with_merged_cells

        path = make_excel_with_merged_cells()
        data = read_excel(str(path))

        sheet = data["资产负债表"]
        assert len(sheet.rows) > 0
        asset_row = next(r for r in sheet.rows if r["项目"] == "资产总计")
        assert asset_row["期末余额"] == 2000000.00
        # 合并单元格的行数据应从第3行开始
        assert asset_row["_row"] == 4


class TestRawSheetDataProperties:
    """测试 RawSheetData 属性。"""

    def test_name_property_returns_sheet_name(self) -> None:
        from tests.importer.conftest import make_balance_sheet_excel

        path = make_balance_sheet_excel()
        data = read_excel(str(path))

        sheet = data["资产负债表"]
        assert sheet.name == "资产负债表"

    def test_rows_with_zero_values(self) -> None:
        from tests.importer.conftest import make_balance_sheet_excel

        path = make_balance_sheet_excel()
        data = read_excel(str(path))

        sheet = data["资产负债表"]
        treasury_row = next(r for r in sheet.rows if r["项目"] == "库存股")
        assert treasury_row["期末余额"] == 0.0

    def test_rows_with_nonexistent_header_returns_none(self) -> None:
        from tests.importer.conftest import make_balance_sheet_excel

        path = make_balance_sheet_excel()
        data = read_excel(str(path))

        sheet = data["资产负债表"]
        row = sheet.rows[0]
        assert row.get("不存在的列") is None


class TestHeaderRowDetection:
    """测试表头行自动定位与多层表头捕获。"""

    def test_header_row_below_title_rows_is_detected(self) -> None:
        from tests.importer.conftest import make_excel_with_merged_cells

        path = make_excel_with_merged_cells()
        data = read_excel(str(path))
        sheet = data["资产负债表"]
        assert sheet.headers[0] == "项目"
        assert sheet.rows[0]["_row"] == 3

    def test_header_row_without_project_column_uses_amount_keywords(self, tmp_path: Path) -> None:
        """无「项目」列时，凭金额列关键词定位表头（如"资 产"表头）。"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "资产负债表"
        ws.cell(row=1, column=1, value="资产负债表")
        headers = ["资   产", "行次", "期末余额", "年初余额"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=2, column=col_idx, value=header)
        ws.cell(row=3, column=1, value="资产总计")
        ws.cell(row=3, column=3, value=2000000.0)
        path = tmp_path / "no_project_header.xlsx"
        wb.save(str(path))

        data = read_excel(str(path))
        sheet = data["资产负债表"]
        assert sheet.headers[0] == "资   产"
        assert "期末余额" in sheet.headers
        assert sheet.rows[0]["资   产"] == "资产总计"

    def test_multi_layer_header_rows_captured(self, tmp_path: Path) -> None:
        """连续多层表头（如权益变动表）被捕获到 header_rows。"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "所有者权益变动表"
        rows = [
            ["项目", "行次", "46174", None, None, "资本公积", "所有者权益合计"],
            [None, None, None, "股本", "其他权益工具", None, None],
            [None, None, None, None, "优先股", None, None],
            ["一、上年年末余额", 1, None, 1000000.0, None, 2300000.0, 1312769.29],
        ]
        for row_idx, row in enumerate(rows, 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        path = tmp_path / "multi_layer_header.xlsx"
        wb.save(str(path))

        data = read_excel(str(path))
        sheet = data["所有者权益变动表"]
        assert sheet.headers[0] == "项目"
        assert len(sheet.header_rows) == 3
        assert sheet.rows[0]["项目"] == "一、上年年末余额"

    def test_duplicate_header_names_are_suffixed(self, tmp_path: Path) -> None:
        """重复列名（如左右两栏的"期末余额"）自动追加序号后缀。"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "资产负债表"
        headers = ["资   产", "期末余额", "负债和所有者权益", "期末余额"]
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        path = tmp_path / "duplicate_header.xlsx"
        wb.save(str(path))

        data = read_excel(str(path))
        sheet = data["资产负债表"]
        assert sheet.headers == ["资   产", "期末余额", "负债和所有者权益", "期末余额#2"]

    def test_numeric_data_row_stops_header_capture(self, tmp_path: Path) -> None:
        """序号列为空但含数值的数据行不应被当作子表头层。"""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "往来重分类明细"
        rows = [
            ["序号", "账面对应往来科目", "账面余额", "重分类后科目", "重分类后金额"],
            [None, "应收账款", 1000.0, "应收账款", 1000.0],
        ]
        for row_idx, row in enumerate(rows, 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        path = tmp_path / "numeric_first_empty.xlsx"
        wb.save(str(path))

        data = read_excel(str(path))
        sheet = data["往来重分类明细"]
        assert len(sheet.header_rows) == 1
        assert sheet.rows[0]["账面对应往来科目"] == "应收账款"


class TestReadXls:
    """测试 .xls 文件的读取（pandas + xlrd 路径）。"""

    def test_read_xls_roundtrip(self, tmp_path: Path) -> None:
        import xlwt

        path = tmp_path / "balance_sheet.xls"
        book = xlwt.Workbook()
        sheet = book.add_sheet("资产负债表")
        headers = ["项目", "行次", "期末余额"]
        for col_idx, header in enumerate(headers):
            sheet.write(0, col_idx, header)
        rows = [
            ["资产总计", 38, 2000000.0],
            ["负债合计", 69, 1000000.0],
            ["所有者权益合计", 82, 1000000.0],
        ]
        for row_idx, row in enumerate(rows, 1):
            for col_idx, value in enumerate(row):
                sheet.write(row_idx, col_idx, value)
        book.save(str(path))

        data = read_excel(str(path))
        sheet = data["资产负债表"]
        assert sheet.headers[:3] == ["项目", "行次", "期末余额"]
        assert len(sheet.rows) == 3
        assert sheet.rows[0]["项目"] == "资产总计"


class TestReadXlsMissingDependency:
    """测试 .xls 读取缺依赖（xlrd/pandas）时走 COM 回退并抛中文 FSAError。"""

    def test_missing_xlrd_import_error_falls_back_to_com(self, monkeypatch: pytest.MonkeyPatch) -> None:
        """缺 xlrd 的 ImportError 应被 COM 回退链路捕获，最终抛中文 FSAError。"""
        import fsa.core.importer.excel_reader as excel_reader

        monkeypatch.setattr(
            excel_reader,
            "_read_xls",
            lambda path: (_ for _ in ()).throw(ImportError("读取 .xls 需要安装 pandas 与 xlrd")),
        )
        monkeypatch.setattr(
            excel_reader,
            "read_excel_com",
            lambda path: (_ for _ in ()).throw(FSAError("未安装 pywin32，无法使用 Excel COM 读取加密文件")),
        )

        with pytest.raises(FSAError) as excinfo:
            read_excel("missing_dep.xls")

        message = str(excinfo.value)
        assert "读取 .xls 需要安装 pandas 与 xlrd" in message
        assert "Excel COM 打开也失败" in message


class TestReadCsv:
    """CSV 读取: 编码探测与 RawSheetData 转换。"""

    def test_read_utf8_sig_csv(self, tmp_path: Path) -> None:
        path = tmp_path / "利润表.csv"
        path.write_bytes("\ufeff项目,本期金额,上期金额\n营业收入,1000,900\n净利润,500,450\n".encode("utf-8"))
        data = read_excel(str(path))
        sheet = data["利润表"]
        assert sheet.headers == ["项目", "本期金额", "上期金额"]
        row = next(r for r in sheet.rows if r["项目"] == "营业收入")
        assert row["本期金额"] == "1000"
        assert row["_row"] == 2

    def test_read_gbk_csv(self, tmp_path: Path) -> None:
        path = tmp_path / "资产负债表.csv"
        path.write_bytes("项目,期末余额,年初余额\n资产总计,2000000,1850000\n负债合计,1000000,900000\n".encode("gbk"))
        data = read_excel(str(path))
        sheet = data["资产负债表"]
        row = next(r for r in sheet.rows if r["项目"] == "资产总计")
        assert row["期末余额"] == "2000000"
        assert row["年初余额"] == "1850000"

    def test_read_unknown_encoding_raises_fsa_error(self, tmp_path: Path) -> None:
        path = tmp_path / "bad.csv"
        path.write_bytes(b"\xff\xfe\x00\x01\xff")
        with pytest.raises(FSAError, match="编码无法识别"):
            read_excel(str(path))


class TestReadXlsm:
    """openpyxl 原生读取 .xlsm 宏工作簿数据部分。"""

    def test_read_xlsm_sheet(self, tmp_path: Path) -> None:
        from openpyxl import Workbook

        path = tmp_path / "带宏报表.xlsm"
        wb = Workbook()
        ws = wb.active
        ws.title = "资产负债表"
        ws.append(["项目", "期末余额"])
        ws.append(["资产总计", 100.0])
        wb.save(path)
        wb.close()

        data = read_excel(str(path))
        assert "资产负债表" in data
        assert data["资产负债表"].headers == ["项目", "期末余额"]
        assert data["资产负债表"].rows[0]["期末余额"] == 100.0
