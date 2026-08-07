"""测试 Excel 文件生成器: 使用 openpyxl 创建模拟财务报表 Excel 文件。

每个工厂函数返回一个临时文件路径，测试结束后自动清理。
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def _active_sheet(wb: openpyxl.Workbook) -> Worksheet:
    """获取工作簿的活动工作表，处理类型窄化。"""
    sheet = wb.active
    assert sheet is not None
    return sheet


def _write_bs_data(sheet: Worksheet) -> None:
    """写入标准资产负债表数据到工作表。"""
    headers = ["项目", "行次", "期末余额", "年初余额"]
    for col_idx, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_idx, value=header)

    rows = [
        ("流动资产合计", 1, 500000.00, 450000.00),
        ("货币资金", 2, 100000.00, 80000.00),
        ("应收账款", 3, 180000.00, 160000.00),
        ("预付款项", 4, 50000.00, 45000.00),
        ("其他应收款", 5, 30000.00, 25000.00),
        ("非流动资产合计", 10, 1500000.00, 1400000.00),
        ("在建工程", 11, 200000.00, 150000.00),
        ("资产总计", 20, 2000000.00, 1850000.00),
        ("流动负债合计", 25, 600000.00, 550000.00),
        ("应付账款", 26, 300000.00, 280000.00),
        ("预收款项", 27, 100000.00, 90000.00),
        ("其他应付款", 28, 50000.00, 45000.00),
        ("非流动负债合计", 30, 400000.00, 350000.00),
        ("负债合计", 35, 1000000.00, 900000.00),
        ("实收资本", 40, 500000.00, 500000.00),
        ("资本公积", 41, 100000.00, 100000.00),
        ("其他综合收益", 42, 50000.00, 40000.00),
        ("盈余公积", 43, 80000.00, 70000.00),
        ("未分配利润", 44, 270000.00, 240000.00),
        ("库存股", 45, 0.00, 0.00),
        ("所有者权益合计", 50, 1000000.00, 950000.00),
    ]
    for row_idx, (name, row_num, ending, beginning) in enumerate(rows, 2):
        sheet.cell(row=row_idx, column=1, value=name)
        sheet.cell(row=row_idx, column=2, value=row_num)
        sheet.cell(row=row_idx, column=3, value=ending)
        sheet.cell(row=row_idx, column=4, value=beginning)


def _write_is_data(sheet: Worksheet) -> None:
    """写入标准利润表数据到工作表。"""
    headers = ["项目", "行次", "本期金额", "上期金额"]
    for col_idx, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_idx, value=header)

    rows = [
        ("营业收入", 1, 3000000.00, 2800000.00),
        ("主营业务收入", 2, 2500000.00, 2400000.00),
        ("其他业务收入", 3, 500000.00, 400000.00),
        ("营业成本", 4, 2000000.00, 1900000.00),
        ("税金及附加", 5, 50000.00, 48000.00),
        ("销售费用", 6, 200000.00, 180000.00),
        ("管理费用", 7, 150000.00, 140000.00),
        ("研发费用", 8, 80000.00, 75000.00),
        ("财务费用", 9, 30000.00, 28000.00),
        ("资产减值损失", 10, 10000.00, 0.00),
        ("信用减值损失", 11, 5000.00, 0.00),
        ("其他收益", 12, 20000.00, 15000.00),
        ("投资收益", 13, 50000.00, 40000.00),
        ("公允价值变动收益", 14, 0.00, 0.00),
        ("资产处置收益", 15, 0.00, 0.00),
        ("营业利润", 20, 545000.00, 484000.00),
        ("营业外收入", 21, 5000.00, 3000.00),
        ("营业外支出", 22, 10000.00, 8000.00),
        ("利润总额", 25, 540000.00, 479000.00),
        ("所得税费用", 26, 135000.00, 119750.00),
        ("净利润", 30, 405000.00, 359250.00),
        ("税后其他综合收益", 31, 10000.00, 5000.00),
        ("综合收益总额", 32, 415000.00, 364250.00),
        ("扣除非经常性损益的净利润", 33, 395000.00, 350000.00),
    ]
    for row_idx, (name, row_num, current, prior) in enumerate(rows, 2):
        sheet.cell(row=row_idx, column=1, value=name)
        sheet.cell(row=row_idx, column=2, value=row_num)
        sheet.cell(row=row_idx, column=3, value=current)
        sheet.cell(row=row_idx, column=4, value=prior)


def _write_cf_data(sheet: Worksheet) -> None:
    """写入标准现金流量表数据到工作表。"""
    headers = ["项目", "行次", "本期金额", "上期金额"]
    for col_idx, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_idx, value=header)

    rows = [
        ("经营活动产生的现金流量净额", 1, 500000.00, 450000.00),
        ("销售商品、提供劳务收到的现金", 2, 3200000.00, 3000000.00),
        ("投资活动产生的现金流量净额", 5, -200000.00, -150000.00),
        ("筹资活动产生的现金流量净额", 10, -100000.00, -80000.00),
        ("汇率变动对现金的影响", 12, 5000.00, 3000.00),
        ("现金及现金等价物净增加额", 15, 205000.00, 223000.00),
        ("期初现金及现金等价物余额", 16, 800000.00, 577000.00),
        ("期末现金及现金等价物余额", 17, 1005000.00, 800000.00),
    ]
    for row_idx, (name, row_num, current, prior) in enumerate(rows, 2):
        sheet.cell(row=row_idx, column=1, value=name)
        sheet.cell(row=row_idx, column=2, value=row_num)
        sheet.cell(row=row_idx, column=3, value=current)
        sheet.cell(row=row_idx, column=4, value=prior)


def make_balance_sheet_excel(tmp_path: Path | None = None) -> Path:
    """创建标准资产负债表 Excel 文件。"""
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())
    filepath = tmp_path / "balance_sheet.xlsx"
    wb = openpyxl.Workbook()
    sheet = _active_sheet(wb)
    sheet.title = "资产负债表"
    _write_bs_data(sheet)
    wb.save(str(filepath))
    return filepath


def make_income_statement_excel(tmp_path: Path | None = None) -> Path:
    """创建标准利润表 Excel 文件。"""
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())
    filepath = tmp_path / "income_statement.xlsx"
    wb = openpyxl.Workbook()
    sheet = _active_sheet(wb)
    sheet.title = "利润表"
    _write_is_data(sheet)
    wb.save(str(filepath))
    return filepath


def make_cash_flow_excel(tmp_path: Path | None = None) -> Path:
    """创建标准现金流量表 Excel 文件。"""
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())
    filepath = tmp_path / "cash_flow.xlsx"
    wb = openpyxl.Workbook()
    sheet = _active_sheet(wb)
    sheet.title = "现金流量表"
    _write_cf_data(sheet)
    wb.save(str(filepath))
    return filepath


def make_multi_sheet_excel(tmp_path: Path | None = None) -> Path:
    """创建包含三大报表的 Excel 文件。"""
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())
    filepath = tmp_path / "multi_sheet.xlsx"
    wb = openpyxl.Workbook()
    ws_bs = _active_sheet(wb)
    ws_bs.title = "资产负债表"
    _write_bs_data(ws_bs)
    ws_is = wb.create_sheet("利润表")
    _write_is_data(ws_is)
    ws_cf = wb.create_sheet("现金流量表")
    _write_cf_data(ws_cf)
    wb.save(str(filepath))
    return filepath


def make_empty_excel(tmp_path: Path | None = None) -> Path:
    """创建空的 Excel 文件（仅含默认空工作表）。"""
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())
    filepath = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(filepath))
    return filepath


def make_excel_with_merged_cells(tmp_path: Path | None = None) -> Path:
    """创建包含合并单元格的资产负债表 Excel 文件。"""
    if tmp_path is None:
        tmp_path = Path(tempfile.mkdtemp())
    filepath = tmp_path / "merged_cells.xlsx"
    wb = openpyxl.Workbook()
    sheet = _active_sheet(wb)
    sheet.title = "资产负债表"
    sheet.merge_cells("A1:D1")
    sheet.cell(row=1, column=1, value="资产负债表（合并）")
    # 写入表头到第2行
    headers = ["项目", "行次", "期末余额", "年初余额"]
    for col_idx, header in enumerate(headers, 1):
        sheet.cell(row=2, column=col_idx, value=header)
    # 写入数据从第3行开始
    rows = [
        ("流动资产合计", 1, 500000.00, 450000.00),
        ("资产总计", 20, 2000000.00, 1850000.00),
        ("负债合计", 35, 1000000.00, 900000.00),
        ("所有者权益合计", 50, 1000000.00, 950000.00),
    ]
    for row_idx, (name, row_num, ending, beginning) in enumerate(rows, 3):
        sheet.cell(row=row_idx, column=1, value=name)
        sheet.cell(row=row_idx, column=2, value=row_num)
        sheet.cell(row=row_idx, column=3, value=ending)
        sheet.cell(row=row_idx, column=4, value=beginning)
    wb.save(str(filepath))
    return filepath