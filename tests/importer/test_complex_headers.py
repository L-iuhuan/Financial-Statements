"""复杂表头与多栏布局回归: 多级表头/期间列/左右双栏/重名列。

每个用例现场生成一个独立工作簿, 覆盖 EXECUTION_PLAN A2 的常见变体。
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from fsa.core.importer.excel_reader import read_excel
from fsa.core.importer.item_extractor import extract_items
from fsa.core.models.report import ReportType

_BS = ReportType.BALANCE_SHEET
_IS = ReportType.INCOME_STATEMENT
_CF = ReportType.CASH_FLOW_STATEMENT


def _build_workbook(
    tmp_path: Path,
    sheet_name: str,
    header_rows: list[list[object]],
    data_rows: list[list[object]],
) -> str:
    path = tmp_path / f"{sheet_name}_{abs(hash(sheet_name))}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in header_rows + data_rows:
        ws.append(row)
    wb.save(str(path))
    wb.close()
    return str(path)


def _extract(tmp_path: Path, sheet_name: str, header_rows, data_rows, report_type):
    path = _build_workbook(tmp_path, sheet_name, header_rows, data_rows)
    raw = read_excel(path)[sheet_name]
    return extract_items(raw, report_type)


def _amounts(items):
    return {item.key: (item.amount, item.beginning_amount) for item in items}


class TestMultiLayerHeaders:
    """>2 层表头 + 合并单元格留白场景。"""

    @pytest.mark.parametrize(
        ("title", "header_rows", "data_rows", "report_type", "key", "expected"),
        [
            # 1-4: 资产负债表, 两层表头 (年份/月度期间)
            (
                "资产负债表",
                [["项目", "期末余额", "年初余额"], [None, "2024年", "2023年"]],
                [["资产总计", 900.0, 860.0]],
                _BS,
                "asset_total",
                (900.0, 860.0),
            ),
            (
                "资产负债表",
                [["项目", "期末余额", "年初余额"], [None, "2024年12月", "2023年12月"]],
                [["负债合计", 400.0, 430.0]],
                _BS,
                "liability_total",
                (400.0, 430.0),
            ),
            (
                "资产负债表",
                [["项目", "期末余额", "年初余额"], [None, "2024年", "2023年"], [None, "本年", "上年"]],
                [["所有者权益合计", 500.0, 430.0]],
                _BS,
                "equity_total",
                (500.0, 430.0),
            ),
            (
                "资产负债表",
                [["项目", "期末余额", "年初余额"], [None, "2024年", "2023年"], [None, "合并数", "母公司数"]],
                [["资产总计", 910.0, 870.0]],
                _BS,
                "asset_total",
                (910.0, 870.0),
            ),
            # 5-8: 利润表, 两层表头
            (
                "利润表",
                [["项目", "本期金额", "上期金额"], [None, "2024年", "2023年"]],
                [["营业收入", 100.0, 90.0]],
                _IS,
                "revenue",
                (100.0, 90.0),
            ),
            (
                "利润表",
                [["项目", "本期金额", "上期金额"], [None, "2024年1-6月", "2023年1-6月"]],
                [["净利润", 60.0, 50.0]],
                _IS,
                "net_profit",
                (60.0, 50.0),
            ),
            (
                "利润表",
                [["项目", "本期金额", "上期金额"], [None, "2024年6月", "2023年6月"]],
                [["营业成本", 40.0, 35.0]],
                _IS,
                "operating_cost",
                (40.0, 35.0),
            ),
            (
                "利润表",
                [["项目", "本期金额", "上期金额"], [None, "46174", "45292"]],
                [["营业利润", 30.0, 28.0]],
                _IS,
                "operating_profit",
                (30.0, 28.0),
            ),
            # 9-12: 现金流量表, 两层表头
            (
                "现金流量表",
                [["项目", "本期金额", "上期金额"], [None, "2024年", "2023年"]],
                [["经营活动产生的现金流量净额", 80.0, 70.0]],
                _CF,
                "operating_net",
                (80.0, 70.0),
            ),
            (
                "现金流量表",
                [["项目", "本期金额", "上期金额"], [None, "2024年1-12月", "2023年1-12月"]],
                [["投资活动产生的现金流量净额", -20.0, -15.0]],
                _CF,
                "investing_net",
                (-20.0, -15.0),
            ),
            (
                "现金流量表",
                [["项目", "本期金额", "上期金额"], [None, "2024年12月", "2023年12月"]],
                [["筹资活动产生的现金流量净额", 10.0, 8.0]],
                _CF,
                "financing_net",
                (10.0, 8.0),
            ),
            (
                "现金流量表",
                [["项目", "本期金额", "上期金额"], [None, "46174", "45292"]],
                [["现金及现金等价物净增加额", 70.0, 63.0]],
                _CF,
                "net_increase_cash",
                (70.0, 63.0),
            ),
            # 13-16: 三层表头, 其中一层为空标签 (合并单元格留白)
            (
                "利润表",
                [["项目", "本期金额", "上期金额"], [None, "2024年", "2023年"], [None, None, None]],
                [["营业收入", 111.0, 99.0]],
                _IS,
                "revenue",
                (111.0, 99.0),
            ),
            (
                "现金流量表",
                [["项目", "本期金额", "上期金额"], [None, "2024年", "2023年"], [None, "本期", "上期"]],
                [["期末现金及现金等价物余额", 200.0, 180.0]],
                _CF,
                "ending_cash_equiv",
                (200.0, 180.0),
            ),
            (
                "资产负债表",
                [["项目", "期末余额", "年初余额"], [None, "2024年", "2023年"], [None, "人民币", "人民币"]],
                [["货币资金", 50.0, 45.0]],
                _BS,
                "monetary_funds",
                (50.0, 45.0),
            ),
            (
                "资产负债表",
                [["项目", "期末余额（万元）", "年初余额"], [None, "2024年", "2023年"]],
                [["资产总计", 200.0, 180.0]],
                _BS,
                "asset_total",
                (2_000_000.0, 1_800_000.0),
            ),
            # 17-20: 无"项目"列名、金额关键词定位表头、纯期间列
            (
                "利润表",
                [[None, "本期金额", "上期金额"], [None, "2024年", "2023年"]],
                [["营业收入", 77.0, 66.0]],
                _IS,
                "revenue",
                (77.0, 66.0),
            ),
            (
                "现金流量表",
                [[None, "本期金额", "上期金额"], [None, "2024年1-6月", "2023年1-6月"]],
                [["经营活动产生的现金流量净额", 55.0, 44.0]],
                _CF,
                "operating_net",
                (55.0, 44.0),
            ),
            (
                "利润表",
                [["项目", "2024年", "2023年"]],
                [["营业收入", 88.0, 77.0]],
                _IS,
                "revenue",
                (88.0, 77.0),
            ),
            (
                "现金流量表",
                [["项目", "46174", "45292"]],
                [["经营活动产生的现金流量净额", 99.0, 88.0]],
                _CF,
                "operating_net",
                (99.0, 88.0),
            ),
        ],
    )
    def test_extracts_expected_item(
        self, tmp_path: Path, title, header_rows, data_rows, report_type, key, expected
    ) -> None:
        items = _extract(tmp_path, title, header_rows, data_rows, report_type)
        amounts = _amounts(items)
        assert key in amounts
        amount, beginning = amounts[key]
        assert (amount, beginning) == pytest.approx(expected)
        # 追溯列必须是实际使用的合并后主列, 不能是占位列
        assert "列" not in items[0].column

    def test_four_layer_header_extracts(self, tmp_path: Path) -> None:
        """四层表头 (系统上限) 纵向合并后仍可提取。"""
        items = _extract(
            tmp_path,
            "利润表",
            [
                ["项目", "本期金额", "上期金额"],
                [None, "2024年", "2023年"],
                [None, "上半年", "上半年"],
                [None, "1-6月", "1-6月"],
            ],
            [["营业收入", 120.0, 110.0]],
            _IS,
        )
        assert _amounts(items)["revenue"] == (120.0, 110.0)


class TestLeftRightBalanceSheet:
    """资产负债表左右双栏: 列错位防护与独立提取。"""

    def test_left_right_duplicate_amount_columns(self, tmp_path: Path) -> None:
        path = _build_workbook(
            tmp_path,
            "资产负债表",
            [["资产", "期末余额", "年初余额", "负债和所有者权益", "期末余额", "年初余额"]],
            [
                ["流动资产合计", 500.0, 450.0, "流动负债合计", 300.0, 280.0],
                ["资产总计", 900.0, 860.0, "负债和所有者权益总计", 900.0, 860.0],
            ],
        )
        raw = read_excel(path)["资产负债表"]
        assert raw.headers[1] == "期末余额"
        assert raw.headers[4] == "期末余额#2"
        items = extract_items(raw, _BS)
        amounts = {item.key: item.amount for item in items}
        assert amounts["current_assets"] == 500.0
        assert amounts["current_liabilities"] == 300.0
        assert amounts["asset_total"] == 900.0
        assert "liability_equity_total" in amounts

    def test_left_right_two_layer_period_headers(self, tmp_path: Path) -> None:
        path = _build_workbook(
            tmp_path,
            "资产负债表",
            [
                ["资产", "期末余额", "年初余额", "负债和股东权益", "期末余额", "年初余额"],
                [None, "2024年", "2023年", None, "2024年", "2023年"],
            ],
            [
                ["资产总计", 900.0, 860.0, "负债合计", 400.0, 430.0],
                ["货币资金", 100.0, 90.0, "所有者权益合计", 500.0, 430.0],
            ],
        )
        raw = read_excel(path)["资产负债表"]
        items = extract_items(raw, _BS)
        amounts = {item.key: (item.amount, item.beginning_amount) for item in items}
        assert amounts["asset_total"] == (900.0, 860.0)
        assert amounts["liability_total"] == (400.0, 430.0)
        assert amounts["equity_total"] == (500.0, 430.0)

    def test_merged_name_column_across_header_rows(self, tmp_path: Path) -> None:
        """项目列纵向合并 (A1:A3), 其余金额列两行表头。"""
        path = tmp_path / "bs_merged_name.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "资产负债表"
        ws.append(["项目", "期末余额", "年初余额"])
        ws.append([None, "2024年", "2023年"])
        ws.append(["资产总计", 900.0, 860.0])
        ws.merge_cells("A1:A2")
        wb.save(str(path))
        wb.close()

        raw = read_excel(str(path))["资产负债表"]
        items = extract_items(raw, _BS)
        assert {item.key for item in items} == {"asset_total"}
        assert items[0].amount == 900.0
        assert items[0].beginning_amount == 860.0


class TestDuplicateAmountColumnNames:
    """同名金额列去重: 不互相覆盖, 提取保持确定性。"""

    def test_duplicate_primary_columns_are_suffixed(self, tmp_path: Path) -> None:
        path = _build_workbook(
            tmp_path,
            "利润表",
            [["项目", "本期金额", "本期金额"]],
            [["营业收入", 100.0, 999.0]],
        )
        raw = read_excel(path)["利润表"]
        assert raw.headers == ["项目", "本期金额", "本期金额#2"]
        # 提取稳定取第一列, 第二列不污染结果
        items = extract_items(raw, _IS)
        assert items[0].amount == 100.0

    def test_four_duplicate_names_still_unique(self, tmp_path: Path) -> None:
        path = _build_workbook(
            tmp_path,
            "现金流量表",
            [["项目", "本期金额", "本期金额", "本期金额"]],
            [["经营活动产生的现金流量净额", 10.0, 20.0, 30.0]],
        )
        raw = read_excel(path)["现金流量表"]
        assert raw.headers == ["项目", "本期金额", "本期金额#2", "本期金额#3"]
