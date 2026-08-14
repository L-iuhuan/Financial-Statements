"""明细导入器测试: 从工作簿识别并解析余额表/序时账/现金流量明细。"""

from __future__ import annotations

import io
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path

import openpyxl
from loguru import logger

from fsa.core.importer.detail_importer import DetailImporter
from fsa.core.models.detail import CashFlowDetailRow, DetailDataset, JournalRow


@contextmanager
def _capture_loguru(level: str = "WARNING") -> Iterator[io.StringIO]:
    """捕获 loguru 日志到 StringIO。

    loguru 默认 stderr handler 在导入时即绑定底层 fd，capsys/caplog 均捕获不到，
    因此测试中挂载一个临时 StringIO sink。
    """
    sink = io.StringIO()
    handler_id = logger.add(sink, level=level, format="{level} {message}", colorize=False)
    try:
        yield sink
    finally:
        logger.remove(handler_id)


def _write_sheet(ws: openpyxl.worksheet.worksheet.Worksheet, headers: list[str], rows: list[list[object]]) -> None:
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)


def _make_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "detail.xlsx"
    wb = openpyxl.Workbook()
    ws_tb = wb.active
    ws_tb.title = "带辅助核算的科目余额表（1-本月）"
    _write_sheet(
        ws_tb,
        ["会计期间", "科目编码", "科目名称", "期初余额借方", "本期发生借方", "期末余额借方", "期末余额贷方"],
        [
            ["2026.01 - 2026.06", "1002", "银行存款", 100.0, 500.0, 600.0, 0.0],
            ["2026.01 - 2026.06", "1012", "其他货币资金", 0.0, 0.0, 100.0, 0.0],
        ],
    )
    ws_journal = wb.create_sheet("序时账（1-本月）")
    _write_sheet(
        ws_journal,
        ["日期", "凭证号数", "上级科目", "科目编码", "科目名称", "摘要", "方向", "金额"],
        [
            ["2026-06-30", "记-0001", "银行存款", "1002", "银行存款", "收款", "贷", 500.0],
            ["2026-06-30", "记-0001", "应收账款", "1122", "应收账款", "收款", "借", 500.0],
        ],
    )
    ws_cf = wb.create_sheet("现金流量表（1-本月）")
    _write_sheet(
        ws_cf,
        ["2026年月", "2026年日", "凭证号", "现金流量项目", "摘要", "方向", "金额"],
        [
            [6, 30, "记-0001", "销售商品、提供劳务收到的现金(01)", "货款", "流入", 500.0],
            [None, None, None, "销售商品、提供劳务收到的现金(01)", "小计", "流入", 500.0],
        ],
    )
    ws_rc = wb.create_sheet("往来重分类明细")
    _write_sheet(
        ws_rc,
        ["序号", "账面对应往来科目（重分类前科目）", "客户/供应商/等", "账面余额",
         "重分类后科目", "重分类后金额", "开票金额", "暂估金额"],
        [
            [1, "应收账款", "深圳某公司", -20.0, "预收账款", 20.0, None, None],
            [2, "应付账款", "供应商A", 1000.0, "应付账款", 1000.0, None, None],
        ],
    )
    ws_purchase = wb.create_sheet("关联方采购")
    _write_sheet(
        ws_purchase,
        ["单位名称（填表单位-购买方）", "对方单位名称", "款项性质", "总采购金额",
         "供应链采购", "其中：结存存货", "主营业务成本", "研发费用"],
        [
            ["杭州杰为科技有限公司", "拓尔微电子股份有限公司", "采购款", 1000.0, None, 1000.0, None, None],
        ],
    )
    ws_sales = wb.create_sheet("销售收入成本明细表")
    _write_sheet(
        ws_sales,
        ["年", "月", "归属主体", "客户名称", "收入类型", "销售收入金额", "销售成本金额",
         "直接材料", "加工费", "直接人工", "制造费", "销售毛利率"],
        [
            [2026, 1, "杭州杰为科技有限公司", "客户A", "主营业务收入", 800.0, 500.0,
             300.0, 100.0, 50.0, 50.0, 0.375],
        ],
    )
    ws_internal = wb.create_sheet("内部现金流量明细表")
    _write_sheet(
        ws_internal,
        ["月份", "统计单位名称", "对方单位名称", "款项性质", "现金流量项目", "发生额"],
        [
            [1, "杭州杰为科技有限公司", "拓尔微电子股份有限公司", "货款",
             "收到的其他与经营活动的现金", 100.0],
        ],
    )
    wb.save(str(path))
    return path


class TestDetailImporter:
    """明细导入器按表头识别工作表并解析数据。"""

    def test_import_parses_three_detail_kinds(self, tmp_path: Path) -> None:
        path = _make_workbook(tmp_path)
        dataset = DetailImporter(period="2026-06").import_file(str(path))

        assert len(dataset.trial_balance) == 2
        assert dataset.trial_balance[0].account_code == "1002"
        assert dataset.trial_balance[0].ending_debit == 600.0

        assert len(dataset.journal) == 2
        assert dataset.journal[0].voucher_no == "记-0001"
        assert dataset.journal[0].direction == "贷"

        assert len(dataset.cash_flow_detail) == 2
        assert dataset.cash_flow_detail[0].project == "销售商品、提供劳务收到的现金(01)"

        assert len(dataset.reclassifications) == 2
        assert dataset.reclassifications[0].original_account == "应收账款"
        assert dataset.reclassifications[0].book_amount == -20.0

        assert len(dataset.related_party_purchases) == 1
        assert dataset.related_party_purchases[0].total_amount == 1000.0
        assert len(dataset.sales_details) == 1
        assert dataset.sales_details[0].revenue_amount == 800.0
        assert len(dataset.internal_cash_flows) == 1
        assert dataset.internal_cash_flows[0].amount == 100.0

    def test_current_month_sheet_goes_to_separate_list(self, tmp_path: Path) -> None:
        path = tmp_path / "detail_current.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "带辅助核算的科目余额表（本月）"
        _write_sheet(
            ws,
            ["科目编码", "科目名称", "期末余额借方"],
            [["1002", "银行存款", 123.0]],
        )
        wb.save(str(path))

        dataset = DetailImporter().import_file(str(path))
        assert dataset.trial_balance == []
        assert len(dataset.trial_balance_current) == 1

    def test_current_month_journal_goes_to_journal_current(
        self, tmp_path: Path
    ) -> None:
        """本月口径序时账进入 journal_current 且有中文 warning 日志。"""
        path = tmp_path / "journal_current.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "序时账（本月）"
        _write_sheet(
            ws,
            ["日期", "凭证号", "科目编码", "科目名称", "摘要", "方向", "金额"],
            [
                ["2026-06-30", "记-0001", "1002", "银行存款", "收款", "贷", 500.0],
                ["2026-06-30", "记-0001", "1122", "应收账款", "收款", "借", 500.0],
            ],
        )
        wb.save(str(path))

        with _capture_loguru() as sink:
            dataset = DetailImporter().import_file(str(path))

        assert dataset.journal == []
        assert len(dataset.journal_current) == 2
        assert dataset.journal_current[0].voucher_no == "记-0001"
        # 中文 warning 说明数据被分流而非静默丢弃
        captured = sink.getvalue()
        assert "WARNING" in captured
        assert "序时账" in captured and "本月" in captured
        assert "journal_current" in captured

    def test_cumulative_journal_still_goes_to_journal(self, tmp_path: Path) -> None:
        """累计口径序时账仍进入 journal。"""
        path = tmp_path / "journal_cumulative.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "序时账（1-本月）"
        _write_sheet(
            ws,
            ["日期", "凭证号", "科目编码", "科目名称", "摘要", "方向", "金额"],
            [["2026-06-30", "记-0001", "1002", "银行存款", "收款", "贷", 500.0]],
        )
        wb.save(str(path))

        dataset = DetailImporter().import_file(str(path))
        assert len(dataset.journal) == 1
        assert dataset.journal_current == []

    def test_current_month_cash_flow_detail_goes_to_current(
        self, tmp_path: Path
    ) -> None:
        """本月口径现金流明细进入 cash_flow_detail_current 且有 warning 日志。"""
        path = tmp_path / "cf_detail_current.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "现金流量明细（本月）"
        _write_sheet(
            ws,
            ["年月", "凭证号", "现金流量项目", "摘要", "方向", "金额"],
            [
                [6, "记-0001", "销售商品、提供劳务收到的现金(01)", "货款", "流入", 500.0],
            ],
        )
        wb.save(str(path))

        with _capture_loguru() as sink:
            dataset = DetailImporter().import_file(str(path))

        assert dataset.cash_flow_detail == []
        assert len(dataset.cash_flow_detail_current) == 1
        captured = sink.getvalue()
        assert "WARNING" in captured
        assert "现金流量明细" in captured and "本月" in captured
        assert "cash_flow_detail_current" in captured


class TestDetailDatasetMergeAndEmpty:
    """测试 DetailDataset 合并与空判断包含新增的本月数据集。"""

    def test_merge_combines_journal_current(self) -> None:
        target = DetailDataset()
        other = DetailDataset(journal_current=[JournalRow(
            date="2026-06-30", voucher_no="记-0001", parent_account="",
            account_code="1002", account_name="银行存款", summary="收款",
            direction="贷", amount=500.0,
        )])
        target.merge(other)
        assert len(target.journal_current) == 1

    def test_merge_combines_cash_flow_detail_current(self) -> None:
        target = DetailDataset()
        other = DetailDataset(cash_flow_detail_current=[CashFlowDetailRow(
            voucher_no="记-0001", project="销售商品收到的现金(01)", summary="货款",
            direction="流入", amount=500.0,
        )])
        target.merge(other)
        assert len(target.cash_flow_detail_current) == 1

    def test_is_empty_false_with_only_current_data(self) -> None:
        dataset = DetailDataset(journal_current=[
            JournalRow(date="", voucher_no="记-0001", parent_account="",
                       account_code="1002", account_name="", summary="",
                       direction="借", amount=1.0)
        ])
        assert dataset.is_empty is False
