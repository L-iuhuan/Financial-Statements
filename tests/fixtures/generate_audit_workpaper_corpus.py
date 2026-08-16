"""生成贴近审计底稿样式的财务报表演练语料。

与 generate_realistic_report.py 的区别:
- 每张工作表带审计底稿表头 (被审计单位/审计截止日/编制人/复核人)
- 主表金额互相勾稽
- 同时生成科目余额表、序时账、现金流量明细
- 用于导入/校验/导出链路回归

用法:
    python tests/fixtures/generate_audit_workpaper_corpus.py [输出目录]
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

DEFAULT_OUTPUT = Path("tests/fixtures/real_reports/审计底稿样式_全套报表.xlsx")

_TITLE_FONT = Font(name="宋体", size=14, bold=True)
_HEADER_FONT = Font(name="宋体", size=10, bold=True)
_DATA_FONT = Font(name="宋体", size=10)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _write_workpaper_header(
    ws,
    *,
    unit_name: str,
    statement_name: str,
    form_code: str,
    date_text: str,
    headers: list[str],
    start_row: int = 1,
) -> int:
    """写入审计底稿式表头, 返回数据起始行。"""
    cols = len(headers)
    # 被审计单位
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=cols)
    cell = ws.cell(row=start_row, column=1, value=f"被审计单位：{unit_name}")
    cell.font = _DATA_FONT
    cell.alignment = _LEFT

    # 报表名称 + 会企编号
    r2 = start_row + 1
    mid = cols // 2
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=mid)
    ws.merge_cells(start_row=r2, start_column=mid + 1, end_row=r2, end_column=cols)
    c1 = ws.cell(row=r2, column=1, value=statement_name)
    c1.font = _TITLE_FONT
    c1.alignment = _CENTER
    c2 = ws.cell(row=r2, column=mid + 1, value=form_code)
    c2.font = _DATA_FONT
    c2.alignment = _CENTER

    # 截止日 + 单位
    r3 = start_row + 2
    ws.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=mid)
    ws.merge_cells(start_row=r3, start_column=mid + 1, end_row=r3, end_column=cols)
    c3 = ws.cell(row=r3, column=1, value=date_text)
    c3.font = _DATA_FONT
    c3.alignment = _CENTER
    c4 = ws.cell(row=r3, column=mid + 1, value="单位：元")
    c4.font = _DATA_FONT
    c4.alignment = _CENTER

    # 编制/复核留痕
    r4 = start_row + 3
    ws.merge_cells(start_row=r4, start_column=1, end_row=r4, end_column=mid)
    ws.merge_cells(start_row=r4, start_column=mid + 1, end_row=r4, end_column=cols)
    c5 = ws.cell(row=r4, column=1, value="编制人：　　　　复核人：　　　　日期：")
    c5.font = _DATA_FONT
    c5.alignment = _LEFT

    # 列标题
    r5 = start_row + 4
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=r5, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    return r5 + 1


def _write_rows(ws, start_row: int, rows: list[tuple[object, ...]]) -> None:
    for offset, values in enumerate(rows):
        row = start_row + offset
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = (
                _HEADER_FONT
                if (isinstance(values[0], str) and any(key in values[0] for key in ("合计", "总计", "净额")))
                else _DATA_FONT
            )
            cell.border = _BORDER
            if col_idx == 1:
                cell.alignment = _LEFT
            elif col_idx == 2:
                cell.alignment = _CENTER
            else:
                cell.alignment = _RIGHT


def _add_balance_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("资产负债表")
    start = _write_workpaper_header(
        ws,
        unit_name="测试科技有限公司",
        statement_name="资产负债表",
        form_code="会企01表",
        date_text="2024年12月31日",
        headers=["项目", "行次", "期末余额", "年初余额"],
    )
    _write_rows(
        ws,
        start,
        [
            ("流动资产：", "", "", ""),
            ("货币资金", 1, 2_000_000.00, 1_800_000.00),
            ("交易性金融资产", 2, 100_000.00, 80_000.00),
            ("应收票据", 3, 200_000.00, 150_000.00),
            ("应收账款", 4, 800_000.00, 750_000.00),
            ("预付款项", 5, 200_000.00, 180_000.00),
            ("其他应收款", 6, 150_000.00, 120_000.00),
            ("存货", 7, 850_000.00, 900_000.00),
            ("流动资产合计", 8, 4_500_000.00, 3_980_000.00),
            ("非流动资产：", "", "", ""),
            ("固定资产", 9, 3_000_000.00, 3_200_000.00),
            ("在建工程", 10, 500_000.00, 400_000.00),
            ("无形资产", 11, 600_000.00, 650_000.00),
            ("商誉", 12, 200_000.00, 200_000.00),
            ("长期待摊费用", 13, 100_000.00, 120_000.00),
            ("递延所得税资产", 14, 100_000.00, 80_000.00),
            ("非流动资产合计", 15, 4_500_000.00, 4_650_000.00),
            ("资产总计", 16, 9_000_000.00, 8_630_000.00),
            ("流动负债：", "", "", ""),
            ("短期借款", 17, 800_000.00, 1_000_000.00),
            ("应付票据", 18, 150_000.00, 200_000.00),
            ("应付账款", 19, 1_200_000.00, 1_100_000.00),
            ("预收款项", 20, 300_000.00, 250_000.00),
            ("应付职工薪酬", 21, 200_000.00, 180_000.00),
            ("应交税费", 22, 150_000.00, 120_000.00),
            ("其他应付款", 23, 200_000.00, 220_000.00),
            ("流动负债合计", 24, 2_500_000.00, 3_070_000.00),
            ("非流动负债：", "", "", ""),
            ("长期借款", 25, 1_200_000.00, 1_000_000.00),
            ("租赁负债", 26, 100_000.00, 120_000.00),
            ("递延所得税负债", 27, 200_000.00, 150_000.00),
            ("非流动负债合计", 28, 1_500_000.00, 1_270_000.00),
            ("负债合计", 29, 4_000_000.00, 4_340_000.00),
            ("实收资本", 30, 3_000_000.00, 3_000_000.00),
            ("资本公积", 31, 500_000.00, 500_000.00),
            ("其他综合收益", 32, 50_000.00, 30_000.00),
            ("盈余公积", 33, 300_000.00, 250_000.00),
            ("未分配利润", 34, 1_200_000.00, 760_000.00),
            ("所有者权益合计", 35, 5_000_000.00, 4_290_000.00),
            ("负债和所有者权益总计", 36, 9_000_000.00, 8_630_000.00),
        ],
    )
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16


def _add_income_statement(wb: Workbook) -> None:
    ws = wb.create_sheet("利润表")
    start = _write_workpaper_header(
        ws,
        unit_name="测试科技有限公司",
        statement_name="利润表",
        form_code="会企02表",
        date_text="2024年度",
        headers=["项目", "行次", "本期金额", "上期金额"],
    )
    _write_rows(
        ws,
        start,
        [
            ("一、营业收入", 1, 5_000_000.00, 4_500_000.00),
            ("减：营业成本", 2, 3_000_000.00, 2_700_000.00),
            ("税金及附加", 3, 100_000.00, 90_000.00),
            ("销售费用", 4, 300_000.00, 280_000.00),
            ("管理费用", 5, 400_000.00, 380_000.00),
            ("研发费用", 6, 200_000.00, 180_000.00),
            ("财务费用", 7, 100_000.00, 120_000.00),
            ("加：其他收益", 8, 20_000.00, 15_000.00),
            ("投资收益", 9, 50_000.00, 40_000.00),
            ("公允价值变动收益", 10, 0.00, 0.00),
            ("信用减值损失", 11, -30_000.00, -20_000.00),
            ("资产减值损失", 12, -50_000.00, -40_000.00),
            ("资产处置收益", 13, 10_000.00, 5_000.00),
            ("二、营业利润", 14, 920_000.00, 760_000.00),
            ("加：营业外收入", 15, 50_000.00, 40_000.00),
            ("减：营业外支出", 16, 30_000.00, 25_000.00),
            ("三、利润总额", 17, 940_000.00, 775_000.00),
            ("减：所得税费用", 18, 235_000.00, 193_750.00),
            ("四、净利润", 19, 705_000.00, 581_250.00),
            ("五、其他综合收益的税后净额", 20, 20_000.00, 15_000.00),
            ("六、综合收益总额", 21, 725_000.00, 596_250.00),
        ],
    )
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16


def _add_cash_flow_statement(wb: Workbook) -> None:
    ws = wb.create_sheet("现金流量表")
    start = _write_workpaper_header(
        ws,
        unit_name="测试科技有限公司",
        statement_name="现金流量表",
        form_code="会企03表",
        date_text="2024年度",
        headers=["项目", "行次", "本期金额", "上期金额"],
    )
    _write_rows(
        ws,
        start,
        [
            ("一、经营活动产生的现金流量：", "", "", ""),
            ("销售商品、提供劳务收到的现金", 1, 4_800_000.00, 4_300_000.00),
            ("收到的税费返还", 2, 50_000.00, 40_000.00),
            ("收到其他与经营活动有关的现金", 3, 100_000.00, 90_000.00),
            ("经营活动现金流入小计", 4, 4_950_000.00, 4_430_000.00),
            ("购买商品、接受劳务支付的现金", 5, 3_200_000.00, 2_900_000.00),
            ("支付给职工以及为职工支付的现金", 6, 500_000.00, 480_000.00),
            ("支付的各项税费", 7, 300_000.00, 280_000.00),
            ("支付其他与经营活动有关的现金", 8, 120_000.00, 110_000.00),
            ("经营活动现金流出小计", 9, 4_120_000.00, 3_770_000.00),
            ("经营活动产生的现金流量净额", 10, 830_000.00, 660_000.00),
            ("二、投资活动产生的现金流量：", "", "", ""),
            ("收回投资收到的现金", 11, 100_000.00, 80_000.00),
            ("取得投资收益收到的现金", 12, 50_000.00, 40_000.00),
            ("处置固定资产、无形资产和其他长期资产收回的现金净额", 13, 10_000.00, 5_000.00),
            ("投资活动现金流入小计", 14, 160_000.00, 125_000.00),
            ("购建固定资产、无形资产和其他长期资产支付的现金", 15, 500_000.00, 450_000.00),
            ("投资支付的现金", 16, 100_000.00, 90_000.00),
            ("投资活动现金流出小计", 17, 600_000.00, 540_000.00),
            ("投资活动产生的现金流量净额", 18, -440_000.00, -415_000.00),
            ("三、筹资活动产生的现金流量：", "", "", ""),
            ("取得借款收到的现金", 19, 800_000.00, 700_000.00),
            ("筹资活动现金流入小计", 20, 800_000.00, 700_000.00),
            ("偿还债务支付的现金", 21, 600_000.00, 500_000.00),
            ("分配股利、利润或偿付利息支付的现金", 22, 100_000.00, 80_000.00),
            ("筹资活动现金流出小计", 23, 700_000.00, 580_000.00),
            ("筹资活动产生的现金流量净额", 24, 100_000.00, 120_000.00),
            ("四、汇率变动对现金及现金等价物的影响", 25, 10_000.00, 5_000.00),
            ("五、现金及现金等价物净增加额", 26, 500_000.00, 370_000.00),
            ("加：期初现金及现金等价物余额", 27, 1_500_000.00, 1_130_000.00),
            ("六、期末现金及现金等价物余额", 28, 2_000_000.00, 1_500_000.00),
        ],
    )
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16


def _add_trial_balance(wb: Workbook) -> None:
    ws = wb.create_sheet("科目余额表（1-本月）")
    start = _write_workpaper_header(
        ws,
        unit_name="测试科技有限公司",
        statement_name="科目余额表",
        form_code="审01表",
        date_text="2024年1-12月",
        headers=[
            "科目编码",
            "科目名称",
            "期初余额借方",
            "期初余额贷方",
            "本期发生借方",
            "本期发生贷方",
            "期末余额借方",
            "期末余额贷方",
        ],
    )
    _write_rows(
        ws,
        start,
        [
            ("1002", "银行存款", 1_500_000.00, 0, 4_950_000.00, 4_450_000.00, 2_000_000.00, 0),
            ("1122", "应收账款", 800_000.00, 0, 0, 0, 800_000.00, 0),
            ("2202", "应付账款", 0, 1_100_000.00, 0, 100_000.00, 0, 1_200_000.00),
        ],
    )
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 18


def _add_journal(wb: Workbook) -> None:
    ws = wb.create_sheet("序时账（1-本月）")
    start = _write_workpaper_header(
        ws,
        unit_name="测试科技有限公司",
        statement_name="序时账",
        form_code="审02表",
        date_text="2024年1-12月",
        headers=["日期", "凭证号", "上级科目", "科目编码", "科目名称", "摘要", "方向", "金额"],
    )
    _write_rows(
        ws,
        start,
        [
            ("2024-06-30", "记-0001", "银行存款", "1002", "银行存款", "收到货款", "借", 500_000.00),
            ("2024-06-30", "记-0001", "应收账款", "1122", "应收账款", "收到货款", "贷", 500_000.00),
        ],
    )
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 16


def _add_cash_flow_detail(wb: Workbook) -> None:
    ws = wb.create_sheet("现金流量明细表（1-本月）")
    start = _write_workpaper_header(
        ws,
        unit_name="测试科技有限公司",
        statement_name="现金流量明细表",
        form_code="审03表",
        date_text="2024年1-12月",
        headers=["凭证号", "现金流量项目", "摘要", "方向", "金额"],
    )
    _write_rows(
        ws,
        start,
        [
            ("记-0001", "销售商品、提供劳务收到的现金", "收到货款", "流入", 500_000.00),
            ("记-0002", "购买商品、接受劳务支付的现金", "支付货款", "流出", 300_000.00),
        ],
    )
    for col in "ABCDE":
        ws.column_dimensions[col].width = 26


def generate(output: Path) -> Path:
    """生成审计底稿样式全套报表工作簿，返回输出路径。"""
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    _add_balance_sheet(wb)
    _add_income_statement(wb)
    _add_cash_flow_statement(wb)
    _add_trial_balance(wb)
    _add_journal(wb)
    _add_cash_flow_detail(wb)
    wb.save(output)
    wb.close()
    return output


def main() -> None:
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUTPUT
    generate(output)
    print(f"已生成: {output}")


if __name__ == "__main__":
    main()
