"""生成上市公司年报样式合并报表语料 (Excel 优先格式变体)。

与审计底稿样式语料的差异:
- 表名带"合并"前缀, 覆盖公开年报常见的 4+1 张主表命名
- 合并资产负债表为左右双栏 (资产 | 负债和所有者权益), 每侧 行次/附注 列
- 金额互相勾稽: BS/IS/CF/SCE 内部及跨表规则全部通过
- 预埋 0 处差异, 用于验证"上市公司格式"导入的正确性基线

用法:
    python tests/fixtures/generate_listed_corpus.py [输出xlsx路径]
"""

from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

DEFAULT_OUTPUT = Path("tests/fixtures/real_reports/上市公司年报样式_合并报表.xlsx")

_TITLE_FONT = Font(name="宋体", size=14, bold=True)
_HEADER_FONT = Font(name="宋体", size=10, bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")


def _write_title_block(
    ws,
    *,
    statement_name: str,
    date_text: str,
    columns: int,
) -> None:
    """写年报式表头: 报表名称 + 编制单位/日期/单位, 返回表头行号。"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    title = ws.cell(row=1, column=1, value=statement_name)
    title.font = _TITLE_FONT
    title.alignment = _CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    meta = ws.cell(
        row=2,
        column=1,
        value=f"编制单位：测试股份公司　　{date_text}　　单位：元",
    )
    meta.alignment = _CENTER


def _write_header(ws, headers: list[str], row: int) -> int:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
    return row + 1


def _write_balance_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("合并资产负债表")
    headers = [
        "资产",
        "行次",
        "附注",
        "期末余额",
        "年初余额",
        "负债和所有者权益（或股东权益）",
        "行次",
        "附注",
        "期末余额",
        "年初余额",
    ]
    _write_title_block(
        ws,
        statement_name="合并资产负债表",
        date_text="2024年12月31日",
        columns=len(headers),
    )
    start = _write_header(ws, headers, row=3)

    rows = [
        ["货币资金", 1, None, 2_000_000.0, 1_800_000.0, None, None, None, None, None],
        ["应收账款", 2, None, 800_000.0, 750_000.0, None, None, None, None, None],
        ["流动资产合计", 3, None, 4_500_000.0, 3_980_000.0, None, None, None, None, None],
        ["固定资产", 4, None, 3_000_000.0, 3_200_000.0, None, None, None, None, None],
        ["非流动资产合计", 5, None, 4_500_000.0, 4_650_000.0, None, None, None, None, None],
        ["资产总计", 6, None, 9_000_000.0, 8_630_000.0, None, None, None, None, None],
        [None, None, None, None, None, "短期借款", 7, None, 800_000.0, 1_000_000.0],
        [None, None, None, None, None, "应付账款", 8, None, 1_200_000.0, 1_100_000.0],
        [None, None, None, None, None, "流动负债合计", 9, None, 2_500_000.0, 3_070_000.0],
        [None, None, None, None, None, "长期借款", 10, None, 1_200_000.0, 1_000_000.0],
        [None, None, None, None, None, "非流动负债合计", 11, None, 1_500_000.0, 1_270_000.0],
        [None, None, None, None, None, "负债合计", 12, None, 4_000_000.0, 4_340_000.0],
        [None, None, None, None, None, "实收资本", 13, None, 3_000_000.0, 3_000_000.0],
        [None, None, None, None, None, "资本公积", 14, None, 500_000.0, 500_000.0],
        [None, None, None, None, None, "盈余公积", 15, None, 300_000.0, 250_000.0],
        [None, None, None, None, None, "未分配利润", 16, None, 700_000.0, 110_000.0],
        [None, None, None, None, None, "归属于母公司所有者权益合计", 17, None, 4_500_000.0, 3_860_000.0],
        [None, None, None, None, None, "少数股东权益", 18, None, 500_000.0, 430_000.0],
        [None, None, None, None, None, "所有者权益合计", 19, None, 5_000_000.0, 4_290_000.0],
    ]
    for offset, values in enumerate(rows):
        for col_idx, value in enumerate(values, 1):
            ws.cell(row=start + offset, column=col_idx, value=value)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["F"].width = 30


def _write_income_statement(wb: Workbook) -> None:
    ws = wb.create_sheet("合并利润表")
    headers = ["项目", "行次", "附注", "本期金额", "上期金额"]
    _write_title_block(ws, statement_name="合并利润表", date_text="2024年度", columns=len(headers))
    start = _write_header(ws, headers, row=3)
    rows = [
        ("一、营业总收入", 1, 5_000_000.0, 4_500_000.0),
        ("其中：营业收入", 2, 5_000_000.0, 4_500_000.0),
        ("减：营业成本", 3, 3_000_000.0, 2_700_000.0),
        ("税金及附加", 4, 100_000.0, 90_000.0),
        ("销售费用", 5, 300_000.0, 280_000.0),
        ("管理费用", 6, 400_000.0, 380_000.0),
        ("研发费用", 7, 200_000.0, 180_000.0),
        ("财务费用", 8, 100_000.0, 120_000.0),
        ("加：其他收益", 9, 20_000.0, 15_000.0),
        ("投资收益", 10, 50_000.0, 40_000.0),
        ("公允价值变动收益", 11, 0.0, 0.0),
        ("信用减值损失", 12, -30_000.0, -20_000.0),
        ("资产减值损失", 13, -50_000.0, -40_000.0),
        ("资产处置收益", 14, 10_000.0, 5_000.0),
        ("二、营业利润", 15, 920_000.0, 760_000.0),
        ("加：营业外收入", 16, 50_000.0, 40_000.0),
        ("减：营业外支出", 17, 30_000.0, 25_000.0),
        ("三、利润总额", 18, 940_000.0, 775_000.0),
        ("减：所得税费用", 19, 235_000.0, 193_750.0),
        ("四、净利润", 20, 705_000.0, 581_250.0),
        ("五、其他综合收益的税后净额", 21, 0.0, 0.0),
        ("六、综合收益总额", 22, 705_000.0, 581_250.0),
    ]
    for offset, (name, row_no, amount, prior) in enumerate(rows):
        ws.cell(row=start + offset, column=1, value=name)
        ws.cell(row=start + offset, column=2, value=row_no)
        ws.cell(row=start + offset, column=4, value=amount)
        ws.cell(row=start + offset, column=5, value=prior)
    ws.column_dimensions["A"].width = 34


def _write_cash_flow(wb: Workbook) -> None:
    ws = wb.create_sheet("合并现金流量表")
    headers = ["项目", "行次", "附注", "本期金额", "上期金额"]
    _write_title_block(ws, statement_name="合并现金流量表", date_text="2024年度", columns=len(headers))
    start = _write_header(ws, headers, row=3)
    rows = [
        ("一、经营活动产生的现金流量：", 0, None, None),
        ("销售商品、提供劳务收到的现金", 1, 4_800_000.0, 4_300_000.0),
        ("收到的税费返还", 2, 50_000.0, 40_000.0),
        ("收到其他与经营活动有关的现金", 3, 100_000.0, 90_000.0),
        ("经营活动现金流入小计", 4, 4_950_000.0, 4_430_000.0),
        ("购买商品、接受劳务支付的现金", 5, 3_200_000.0, 2_900_000.0),
        ("支付给职工以及为职工支付的现金", 6, 500_000.0, 480_000.0),
        ("支付的各项税费", 7, 300_000.0, 280_000.0),
        ("支付其他与经营活动有关的现金", 8, 120_000.0, 110_000.0),
        ("经营活动现金流出小计", 9, 4_120_000.0, 3_770_000.0),
        ("经营活动产生的现金流量净额", 10, 830_000.0, 660_000.0),
        ("二、投资活动产生的现金流量：", 0, None, None),
        ("收回投资收到的现金", 11, 100_000.0, 80_000.0),
        ("投资活动产生的现金流量净额", 12, -440_000.0, -415_000.0),
        ("三、筹资活动产生的现金流量：", 0, None, None),
        ("取得借款收到的现金", 13, 800_000.0, 700_000.0),
        ("筹资活动产生的现金流量净额", 14, 100_000.0, 120_000.0),
        ("四、汇率变动对现金及现金等价物的影响", 15, 10_000.0, 5_000.0),
        ("五、现金及现金等价物净增加额", 16, 500_000.0, 370_000.0),
        ("加：期初现金及现金等价物余额", 17, 1_500_000.0, 1_130_000.0),
        ("六、期末现金及现金等价物余额", 18, 2_000_000.0, 1_500_000.0),
    ]
    for offset, (name, row_no, amount, prior) in enumerate(rows):
        ws.cell(row=start + offset, column=1, value=name)
        ws.cell(row=start + offset, column=2, value=row_no)
        if amount is not None:
            ws.cell(row=start + offset, column=4, value=amount)
        if prior is not None:
            ws.cell(row=start + offset, column=5, value=prior)
    ws.column_dimensions["A"].width = 48


def _write_sce(wb: Workbook) -> None:
    ws = wb.create_sheet("合并所有者权益变动表")
    headers = [
        "项目",
        "实收资本",
        "资本公积",
        "减：库存股",
        "其他综合收益",
        "盈余公积",
        "未分配利润",
        "少数股东权益",
        "所有者权益合计",
    ]
    _write_title_block(
        ws,
        statement_name="合并所有者权益变动表",
        date_text="2024年度",
        columns=len(headers),
    )
    start = _write_header(ws, headers, row=3)
    rows = [
        ("一、上年年末余额", [3_000_000.0, 500_000.0, 0.0, 0.0, 250_000.0, 110_000.0, 430_000.0, 4_290_000.0]),
        ("二、本年年初余额", [3_000_000.0, 500_000.0, 0.0, 0.0, 250_000.0, 110_000.0, 430_000.0, 4_290_000.0]),
        ("三、本年增减变动金额", [0.0, 0.0, 0.0, 0.0, 50_000.0, 590_000.0, 70_000.0, 710_000.0]),
        ("（一）综合收益总额", [0.0, 0.0, 0.0, 0.0, 0.0, 705_000.0, 0.0, 705_000.0]),
        ("（二）所有者投入和减少资本", [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 70_000.0, 70_000.0]),
        ("（三）利润分配", [0.0, 0.0, 0.0, 0.0, 50_000.0, -115_000.0, 0.0, -65_000.0]),
        ("四、本年年末余额", [3_000_000.0, 500_000.0, 0.0, 0.0, 300_000.0, 700_000.0, 500_000.0, 5_000_000.0]),
    ]
    for offset, (label, values) in enumerate(rows):
        ws.cell(row=start + offset, column=1, value=label)
        for col_idx, value in enumerate(values, 2):
            ws.cell(row=start + offset, column=col_idx, value=value)
    ws.column_dimensions["A"].width = 30


def generate(output: Path) -> Path:
    """生成上市公司年报样式合并报表工作簿, 返回输出路径。"""
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    _write_balance_sheet(wb)
    _write_income_statement(wb)
    _write_cash_flow(wb)
    _write_sce(wb)
    wb.save(output)
    wb.close()
    return output


def main() -> None:
    output = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_OUTPUT
    generate(output)
    print(f"已生成: {output}")


if __name__ == "__main__":
    main()
