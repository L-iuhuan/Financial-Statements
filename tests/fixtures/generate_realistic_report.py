"""生成模拟真实企业财报 Excel 文件。

模拟真实场景:
- 三表合一工作簿 (不同 sheet 是不同报表)
- 多行表头 + 合并单元格 (编制单位/报表名/会企0X表/日期/单位)
- 行次列
- 非标准科目名 (企业常用别名)
- 期初+期末 / 本期+上期 双列金额
- 空行、备注行、空金额
- sheet 名带年份后缀
- 列名变体 ("期末数" vs "期末余额", "本期数" vs "本期金额")
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

OUTPUT_PATH = "tests/fixtures/realistic_report.xlsx"

# 样式
_TITLE_FONT = Font(name="宋体", size=14, bold=True)
_HEADER_FONT = Font(name="宋体", size=10, bold=True)
_DATA_FONT = Font(name="宋体", size=10)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _merge_header(
    ws, title: str, form_code: str, date_str: str,
    headers: list[str], start_row: int = 1,
) -> int:
    """写入多行合并表头, 返回数据起始行。"""
    # 行1: 编制单位 (合并)
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row, end_column=len(headers),
    )
    cell = ws.cell(row=start_row, column=1, value="编制单位: 测试科技有限公司")
    cell.font = _DATA_FONT
    cell.alignment = _LEFT

    # 行2: 报表名 + 会企0X表
    r2 = start_row + 1
    mid = len(headers) // 2
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=mid)
    ws.merge_cells(start_row=r2, start_column=mid + 1, end_row=r2, end_column=len(headers))
    c1 = ws.cell(row=r2, column=1, value=title)
    c1.font = _TITLE_FONT
    c1.alignment = _CENTER
    c2 = ws.cell(row=r2, column=mid + 1, value=form_code)
    c2.font = _DATA_FONT
    c2.alignment = _CENTER

    # 行3: 日期 + 单位
    r3 = start_row + 2
    ws.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=mid)
    ws.merge_cells(start_row=r3, start_column=mid + 1, end_row=r3, end_column=len(headers))
    c3 = ws.cell(row=r3, column=1, value=date_str)
    c3.font = _DATA_FONT
    c3.alignment = _CENTER
    c4 = ws.cell(row=r3, column=mid + 1, value="单位: 元")
    c4.font = _DATA_FONT
    c4.alignment = _CENTER

    # 行4: 列标题
    r4 = start_row + 3
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=r4, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER

    return r4 + 1  # 数据从下一行开始


def _write_row(
    ws, row: int, values: list, bold: bool = False, indent: int = 0,
) -> None:
    """写入一行数据。"""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = _HEADER_FONT if bold else _DATA_FONT
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
        elif col_idx == 2:
            cell.alignment = _CENTER
        else:
            cell.alignment = _RIGHT
        cell.border = _THIN_BORDER


def _write_balance_sheet(ws) -> None:
    """资产负债表: 期末数 + 年初数, 含非标准科目名。"""
    data_start = _merge_header(
        ws, "资产负债表", "会企01表",
        "2024年12月31日",
        ["项目", "行次", "期末数", "年初数"],
    )

    rows = [
        # 流动资产 (使用企业常用别名)
        ("流动资产：", "", "", ""),
        ("  现金及银行存款", 1, 2000000.00, 1800000.00),  # 别名: 货币资金
        ("  交易性金融资产", 2, 100000.00, 80000.00),
        ("  应收票据", 3, 200000.00, 150000.00),
        ("  应收账款净额", 4, 800000.00, 750000.00),  # 别名: 应收账款
        ("  预付账款", 5, 200000.00, 180000.00),  # 别名: 预付款项
        ("  其他应收款", 6, 150000.00, 120000.00),
        ("  存货", 7, 850000.00, 900000.00),
        ("  合同资产", 8, None, None),  # 空值
        ("  持有待售资产", 9, None, None),
        ("流动资产合计", 10, 4500000.00, 3980000.00),
        # 非流动资产
        ("非流动资产：", "", "", ""),
        ("  固定资产净值", 11, 3000000.00, 3200000.00),  # 别名: 固定资产
        ("  在建工程", 12, 500000.00, 400000.00),
        ("  无形资产", 13, 600000.00, 650000.00),
        ("  商誉", 14, 200000.00, 200000.00),
        ("  长期待摊费用", 15, 100000.00, 120000.00),
        ("  递延所得税资产", 16, 100000.00, 80000.00),
        ("非流动资产合计", 17, 4500000.00, 4650000.00),
        ("资产总计", 18, 9000000.00, 8630000.00),
        # 流动负债
        ("流动负债：", "", "", ""),
        ("  短期借款", 19, 800000.00, 1000000.00),
        ("  应付票据", 20, 150000.00, 200000.00),
        ("  应付账款", 21, 1200000.00, 1100000.00),
        ("  预收账款", 22, 300000.00, 250000.00),  # 别名: 预收款项
        ("  应付职工薪酬", 23, 200000.00, 180000.00),
        ("  应交税费", 24, 150000.00, 120000.00),
        ("  其他应付款", 25, 200000.00, 220000.00),
        ("  一年内到期的非流动负债", 26, None, None),
        ("流动负债合计", 27, 2500000.00, 3070000.00),
        # 非流动负债
        ("非流动负债：", "", "", ""),
        ("  长期借款", 28, 1200000.00, 1000000.00),
        ("  应付债券", 29, None, None),
        ("  租赁负债", 30, 100000.00, 120000.00),
        ("  递延所得税负债", 31, 200000.00, 150000.00),
        ("非流动负债合计", 32, 1500000.00, 1270000.00),
        ("负债合计", 33, 4000000.00, 4340000.00),
        # 所有者权益
        ("所有者权益：", "", "", ""),
        ("  股本", 34, 3000000.00, 3000000.00),  # 别名: 实收资本
        ("  资本公积", 35, 500000.00, 500000.00),
        ("  减：库存股", 36, None, None),
        ("  其他综合收益", 37, 50000.00, 30000.00),
        ("  盈余公积", 38, 300000.00, 250000.00),
        ("  未分配利润", 39, 1200000.00, 760000.00),  # 760000 = 760000(期初)
        ("所有者权益合计", 40, 5000000.00, 4290000.00),
        ("负债和所有者权益总计", 41, 9000000.00, 8630000.00),
        # 备注行
        ("", "", "", ""),
        ("注: 以上数据未经审计", "", "", ""),
    ]

    for i, row_data in enumerate(rows):
        is_total = any(
            keyword in row_data[0]
            for keyword in ["合计", "总计"]
        ) if row_data[0] else False
        indent = 0 if (is_total or not row_data[0].strip()) else 0
        _write_row(ws, data_start + i, list(row_data), bold=is_total, indent=indent)

    # 列宽
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16


def _write_income_statement(ws) -> None:
    """利润表: 本期数 + 上期数, 使用本期金额变体列名。"""
    data_start = _merge_header(
        ws, "利润表", "会企02表",
        "2024年度",
        ["项目", "行次", "本期数", "上期数"],
    )

    rows = [
        ("一、营业收入", 1, 5000000.00, 4500000.00),
        ("    减：营业成本", 2, 3000000.00, 2700000.00),
        ("        税金及附加", 3, 100000.00, 90000.00),
        ("        销售费用", 4, 300000.00, 280000.00),
        ("        管理费用", 5, 400000.00, 380000.00),
        ("        研发费用", 6, 200000.00, 180000.00),
        ("        财务费用", 7, 100000.00, 120000.00),
        ("        加：其他收益", 8, 20000.00, 15000.00),
        ("        投资收益", 9, 50000.00, 40000.00),
        ("        公允价值变动收益", 10, 0.00, 0.00),
        ("        信用减值损失", 11, -30000.00, -20000.00),  # 损失为负
        ("        资产减值损失", 12, -50000.00, -40000.00),
        ("        资产处置收益", 13, 10000.00, 5000.00),
        ("二、营业利润", 14, 920000.00, 760000.00),
        ("    加：营业外收入", 15, 50000.00, 40000.00),
        ("    减：营业外支出", 16, 30000.00, 25000.00),
        ("三、利润总额", 17, 940000.00, 775000.00),
        ("    减：所得税费用", 18, 235000.00, 193750.00),
        ("四、净利润", 19, 705000.00, 581250.00),
        ("", "", "", ""),
        ("五、其他综合收益的税后净额", 20, 20000.00, 15000.00),
        ("六、综合收益总额", 21, 725000.00, 596250.00),
        ("", "", "", ""),
        ("注: 本期数据按照企业会计准则编制", "", "", ""),
    ]

    for i, row_data in enumerate(rows):
        is_total = "利润" in row_data[0] or "收益总额" in row_data[0] if row_data[0] else False
        _write_row(ws, data_start + i, list(row_data), bold=is_total)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16


def _write_cash_flow(ws) -> None:
    """现金流量表: 本期金额 + 上期金额。"""
    data_start = _merge_header(
        ws, "现金流量表", "会企03表",
        "2024年度",
        ["项目", "行次", "本期金额", "上期金额"],
    )

    rows = [
        ("一、经营活动产生的现金流量：", "", "", ""),
        ("  销售商品、提供劳务收到的现金", 1, 4500000.00, 4000000.00),
        ("  收到的税费返还", 2, 50000.00, 40000.00),
        ("  收到其他与经营活动有关的现金", 3, 30000.00, 20000.00),
        ("    经营活动现金流入小计", 4, 4580000.00, 4060000.00),
        ("  购买商品、接受劳务支付的现金", 5, 2800000.00, 2500000.00),
        ("  支付给职工以及为职工支付的现金", 6, 600000.00, 550000.00),
        ("  支付的各项税费", 7, 200000.00, 180000.00),
        ("  支付其他与经营活动有关的现金", 8, 150000.00, 130000.00),
        ("    经营活动现金流出小计", 9, 3750000.00, 3360000.00),
        ("  经营活动产生的现金流量净额", 10, 830000.00, 700000.00),
        ("二、投资活动产生的现金流量：", "", "", ""),
        ("  收回投资收到的现金", 11, None, None),
        ("  取得投资收益收到的现金", 12, 50000.00, 40000.00),
        ("  处置固定资产收回的现金净额", 13, 100000.00, 80000.00),
        ("    投资活动现金流入小计", 14, 150000.00, 120000.00),
        ("  购建固定资产支付的现金", 15, 400000.00, 350000.00),
        ("  投资支付的现金", 16, None, None),
        ("    投资活动现金流出小计", 17, 400000.00, 350000.00),
        ("  投资活动产生的现金流量净额", 18, -250000.00, -230000.00),
        ("三、筹资活动产生的现金流量：", "", "", ""),
        ("  取得借款收到的现金", 19, 1000000.00, 800000.00),
        ("    筹资活动现金流入小计", 20, 1000000.00, 800000.00),
        ("  偿还债务支付的现金", 20, 800000.00, 900000.00),
        ("  分配股利支付现金", 21, 100000.00, 80000.00),
        ("    筹资活动现金流出小计", 22, 900000.00, 980000.00),
        ("  筹资活动产生的现金流量净额", 23, 100000.00, -180000.00),
        ("四、现金及现金等价物净增加额", 24, 680000.00, 290000.00),
        ("  加：期初现金及现金等价物余额", 25, 1320000.00, 1030000.00),
        ("五、期末现金及现金等价物余额", 26, 2000000.00, 1320000.00),
        ("", "", "", ""),
        ("补充资料：", "", "", ""),
        ("1. 将净利润调节为经营活动现金流量：", "", "", ""),
        ("  净利润", 27, 705000.00, 581250.00),
        ("  加：资产减值准备", 28, 50000.00, 40000.00),
        ("  固定资产折旧", 29, 200000.00, 180000.00),
        ("  无形资产摊销", 30, 50000.00, 45000.00),
        ("  财务费用", 31, 100000.00, 120000.00),
        ("  存货的减少", 32, -50000.00, 0.00),
        ("  经营性应收项目的减少", 33, -150000.00, -100000.00),
        ("  经营性应付项目的增加", 34, 125000.00, 100000.00),
        ("  经营活动产生的现金流量净额", 35, 830000.00, 700000.00),
    ]

    for i, row_data in enumerate(rows):
        is_total = any(
            keyword in row_data[0]
            for keyword in ["净额", "小计", "净增加额", "余额"]
        ) if row_data[0] else False
        _write_row(ws, data_start + i, list(row_data), bold=is_total)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16


def generate() -> str:
    """生成模拟真实企业财报 Excel, 返回文件路径。"""
    wb = Workbook()
    # 删除默认空 sheet, 统一用 create_sheet
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    # Sheet1: 资产负债表 (sheet名带年份后缀)
    ws1 = wb.create_sheet("资产负债表(2024年度)")
    _write_balance_sheet(ws1)

    # Sheet2: 利润表 (sheet名带空格+年份)
    ws2 = wb.create_sheet("利润表 2024")
    _write_income_statement(ws2)

    # Sheet3: 现金流量表 (标准sheet名)
    ws3 = wb.create_sheet("现金流量表")
    _write_cash_flow(ws3)

    wb.save(OUTPUT_PATH)
    print(f"已生成: {OUTPUT_PATH}")
    return OUTPUT_PATH


if __name__ == "__main__":
    generate()
