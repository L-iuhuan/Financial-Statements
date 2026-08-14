"""GUI 报表包导入与合并校验测试（多文件 + 明细数据）。"""

from __future__ import annotations

from pathlib import Path

import openpyxl


def _make_detail_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "detail.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "序时账（1-本月）"
    headers = ["日期", "凭证号数", "上级科目", "科目编码", "科目名称", "摘要", "方向", "金额"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    ws.cell(row=2, column=1, value="2026-06-30")
    ws.cell(row=2, column=2, value="记-0001")
    ws.cell(row=2, column=4, value="1002")
    ws.cell(row=2, column=7, value="借")
    ws.cell(row=2, column=8, value=100.0)
    wb.save(str(path))
    return path


class TestPackageImport:
    """一次导入主表与明细文件并执行合并校验。"""

    def test_import_files_then_validate(self, qapp, qtbot, app_state, tmp_path) -> None:
        from fsa.gui.main_window import MainWindow
        from tests.importer.conftest import make_multi_sheet_excel

        main = make_multi_sheet_excel(tmp_path)
        detail = _make_detail_workbook(tmp_path)
        window = MainWindow(app_state)
        qtbot.addWidget(window)

        window._import_page._on_files([str(main), str(detail)])

        assert len(app_state.reports) == 3
        assert app_state.detail_dataset is not None
        assert len(app_state.detail_dataset.journal) == 1

        window._import_page.trigger_validate()
        assert app_state.results is not None
        assert app_state.results.total > 0
        window.close()
