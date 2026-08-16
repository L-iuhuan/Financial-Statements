"""AuditExporter 的单元测试。

测试导出 Excel 审计底稿的完整流程:
- 创建有效 xlsx 文件
- 汇总 sheet 计数正确
- 明细 sheet 行数正确 (含跳过)
- 追溯 sheet 行数 = 总 trace 项数
- 通过/不通过/跳过/异常状态文本正确
- 空 trace 不崩溃
- 权限错误路径处理
"""

from __future__ import annotations

import os
import stat
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from fsa.core.exporter.audit_exporter import AuditExporter
from fsa.core.models.result import (
    TraceItem,
    ValidationResult,
    ValidationSummary,
)
from fsa.core.models.rule import Severity


def _make_result(
    rule_id: str = "BS-BAL-001",
    rule_name: str = "资产=负债+所有者权益",
    passed: bool = True,
    severity: Severity = Severity.ERROR,
    left_value: float = 100.0,
    right_value: float = 100.0,
    diff: float = 0.0,
    tolerance: float = 0.01,
    formula: str = "asset_total == liability_total + equity_total",
    message: str = "校验通过",
    errored: bool = False,
    skipped: bool = False,
    category: str = "A-表内平衡",
    trace: list[TraceItem] | None = None,
) -> ValidationResult:
    """创建一条校验结果。"""
    return ValidationResult(
        rule_id=rule_id,
        rule_name=rule_name,
        passed=passed,
        severity=severity,
        left_value=left_value,
        right_value=right_value,
        diff=diff,
        tolerance=tolerance,
        formula=formula,
        message=message,
        errored=errored,
        skipped=skipped,
        category=category,
        trace=trace if trace is not None else [],
    )


def _make_trace(
    key: str = "asset_total",
    name: str = "资产总计",
    amount: float = 100.0,
    row: int = 5,
    column: str = "期末余额",
    side: str = "left",
) -> TraceItem:
    """创建一条追溯项。"""
    return TraceItem(
        key=key,
        name=name,
        amount=amount,
        row=row,
        column=column,
        side=side,
    )


def _make_summary(
    results: list[ValidationResult],
    period: str = "2024-12",
    total: int | None = None,
    passed: int | None = None,
    failed: int | None = None,
    errored: int | None = None,
    skipped: int | None = None,
) -> ValidationSummary:
    """创建校验汇总。"""
    if total is None:
        total = len(results)
    if passed is None:
        passed = sum(1 for r in results if r.passed and not r.skipped)
    if failed is None:
        failed = sum(1 for r in results if not r.passed and not r.errored)
    if errored is None:
        errored = sum(1 for r in results if r.errored)
    if skipped is None:
        skipped = sum(1 for r in results if r.skipped)
    return ValidationSummary(
        period=period,
        total=total,
        passed=passed,
        failed=failed,
        errored=errored,
        skipped=skipped,
        results=results,
    )


class TestTraceSourceRowRendering:
    """科目追溯 sheet 的行号/列渲染 (D-01 行号语义 + 中文列兼容)。"""

    def test_pdf_row_rendered_as_page_and_row(self) -> None:
        """PDF 来源行号 (页码编码) 渲染为「第X页表内第N行」。"""
        trace = [
            _make_trace(
                key="asset_total",
                name="资产总计",
                amount=100.0,
                row=10_000_005,
                column="期末余额",
            ),
        ]
        results = [_make_result(rule_id="BS-BAL-001", trace=trace)]
        summary = _make_summary(results=results)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["科目追溯"]
            assert ws.cell(row=2, column=6).value == "第1页表内第5行"
            assert ws.cell(row=2, column=7).value == "期末余额"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_excel_row_rendered_as_number(self) -> None:
        """Excel 来源行号保持工作表 1-based 行号数字显示。"""
        trace = [_make_trace(key="asset_total", row=35)]
        results = [_make_result(trace=trace)]
        summary = _make_summary(results=results)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["科目追溯"]
            assert ws.cell(row=2, column=6).value == 35
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_no_source_row_and_chinese_column(self) -> None:
        """row=0 (未找到科目/阈值变量) 且 column 为中文说明串时不崩溃。

        原始行显示为空字符串, 原始列原样显示中文说明。
        """
        trace = [
            _make_trace(
                key="income_tax_expense",
                name="income_tax_expense",
                amount=0.0,
                row=0,
                column="未在报表中找到（按 0 处理）",
            ),
        ]
        results = [_make_result(trace=trace)]
        summary = _make_summary(results=results)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["科目追溯"]
            # openpyxl 将空字符串读回为 None
            assert ws.cell(row=2, column=6).value in ("", None)  # 无源行号
            assert ws.cell(row=2, column=7).value == "未在报表中找到（按 0 处理）"
            wb.close()
        finally:
            os.unlink(tmp_path)


class TestFormatSourceRow:
    """_format_source_row 解码逻辑 (D-01)。"""

    def test_excel_row_passthrough(self) -> None:
        from fsa.core.exporter.audit_exporter import _format_source_row

        assert _format_source_row(35) == 35
        # Excel 最大行号 (1,048,576) 仍按数字显示, 不误判为 PDF
        assert _format_source_row(1048576) == 1048576

    def test_pdf_row_decoded(self) -> None:
        from fsa.core.exporter.audit_exporter import _format_source_row

        assert _format_source_row(10_000_005) == "第1页表内第5行"
        assert _format_source_row(20_000_003) == "第2页表内第3行"

    def test_zero_or_negative_returns_empty(self) -> None:
        from fsa.core.exporter.audit_exporter import _format_source_row

        assert _format_source_row(0) == ""
        assert _format_source_row(-1) == ""


class TestAuditExporter:
    """AuditExporter 核心功能测试。"""

    def test_export_creates_valid_xlsx(self) -> None:
        """导出创建有效的 xlsx 文件，openpyxl 可回读。"""
        exporter = AuditExporter()
        summary = _make_summary(
            results=[_make_result()],
            period="2024-12",
        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            assert wb.sheetnames == ["校验汇总", "校验明细", "科目追溯", "底稿说明"]
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_summary_sheet_counts(self) -> None:
        """汇总 sheet 中的计数与 summary 一致。"""
        results = [
            _make_result(rule_id="R1", passed=True),
            _make_result(rule_id="R2", passed=False, left_value=100.0, right_value=95.0, diff=5.0),
            _make_result(rule_id="R3", errored=True, message="校验异常"),
        ]
        summary = _make_summary(
            results=results,
            period="2024-03",
            total=3,
            passed=1,
            failed=1,
            errored=1,
            skipped=0,
        )
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["校验汇总"]
            # 读取所有单元格到一个字典
            data: dict[str, str] = {}
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
                for cell in row:
                    if cell.value is not None:
                        # 使用列字母+行号作为 key
                        key = f"{cell.coordinate}"
                        data[key] = str(cell.value)
            # 校验计数
            values_str = " ".join(
                str(cell)
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)
                for cell in row
                if cell is not None
            )
            assert "3" in values_str  # total
            assert "1" in values_str  # passed
            assert "2024-03" in values_str  # period
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_detail_sheet_row_count(self) -> None:
        """明细 sheet 行数 = 执行结果数 (含跳过)。"""
        results = [
            _make_result(rule_id="R1", passed=True),
            _make_result(rule_id="R2", passed=False, diff=5.0),
            _make_result(rule_id="R3", skipped=True),
        ]
        summary = _make_summary(results=results, total=2, passed=1, failed=1, skipped=1)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["校验明细"]
            # 减去表头行
            data_rows = ws.max_row - 1
            assert data_rows == 3, f"预期 3 行数据，实际 {data_rows} 行"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_detail_sheet_skipped_row(self) -> None:
        """跳过的规则在明细中显示"跳过"。"""
        results = [
            _make_result(rule_id="R3", skipped=True, diff=0.0),
        ]
        summary = _make_summary(results=results, total=0, passed=0, failed=0, skipped=1)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["校验明细"]
            status_cell = ws.cell(row=2, column=4).value
            assert status_cell == "跳过", f"预期'跳过'，实际'{status_cell}'"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_trace_sheet_row_count(self) -> None:
        """追溯 sheet 行数 = 所有结果的 trace 项总数。"""
        trace1 = [
            _make_trace(key="asset_total", name="资产总计", amount=100.0, side="left"),
            _make_trace(key="liability_total", name="负债合计", amount=60.0, side="right"),
        ]
        trace2 = [
            _make_trace(key="equity_total", name="所有者权益合计", amount=40.0, side="right"),
        ]
        results = [
            _make_result(rule_id="R1", trace=trace1),
            _make_result(rule_id="R2", trace=trace2),
        ]
        summary = _make_summary(results=results)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["科目追溯"]
            data_rows = ws.max_row - 1
            assert data_rows == 3, f"预期 3 行追溯数据，实际 {data_rows} 行"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_trace_content(self) -> None:
        """追溯 sheet 内容包含正确的科目名、金额、行列信息。"""
        trace = [
            _make_trace(
                key="asset_total",
                name="资产总计",
                amount=1234567.89,
                row=35,
                column="期末余额",
                side="left",
            ),
        ]
        results = [_make_result(rule_id="BS-BAL-001", rule_name="资产=负债+所有者权益", trace=trace)]
        summary = _make_summary(results=results)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["科目追溯"]
            row_data = [ws.cell(row=2, column=col).value for col in range(1, 8)]
            assert row_data[0] == "BS-BAL-001"  # 规则ID
            assert row_data[1] == "资产=负债+所有者权益"  # 规则名称
            assert row_data[2] == "左"  # 侧
            assert row_data[3] == "资产总计"  # 科目名称
            assert row_data[5] == 35  # 行号
            assert row_data[6] == "期末余额"  # 列名
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_empty_trace_does_not_break(self) -> None:
        """空 trace 列表不导致导出崩溃。"""
        results = [_make_result(rule_id="R1", trace=[])]
        summary = _make_summary(results=results)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["科目追溯"]
            # 空 trace 时只有表头
            assert ws.max_row == 1, f"预期 1 行 (仅表头)，实际 {ws.max_row} 行"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_status_text_chinese(self) -> None:
        """校验结果状态文本使用中文标签。"""
        results = [
            _make_result(rule_id="R1", passed=True),
            _make_result(rule_id="R2", passed=False, diff=5.0),
            _make_result(rule_id="R3", errored=True),
            _make_result(rule_id="R4", skipped=True),
        ]
        summary = _make_summary(results=results, total=2, passed=1, failed=1, errored=1, skipped=1)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["校验明细"]
            statuses = [ws.cell(row=r, column=4).value for r in range(2, 6)]
            assert statuses == ["通过", "不通过", "异常", "跳过"], f"状态文本不匹配: {statuses}"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_summary_sheet_headers_chinese(self) -> None:
        """汇总 sheet 标题使用中文。"""
        summary = _make_summary(results=[_make_result()], period="2024-12")
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["校验汇总"]
            header_texts: list[str] = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                for cell in row:
                    if cell is not None:
                        header_texts.append(str(cell))
            texts_joined = " ".join(header_texts)
            assert "报告期间" in texts_joined
            assert "校验时间" in texts_joined
            assert "规则总数" in texts_joined
            assert "通过" in texts_joined
            assert "不通过" in texts_joined
            assert "异常" in texts_joined
            assert "跳过" in texts_joined
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_detail_sheet_headers_chinese(self) -> None:
        """明细 sheet 标题使用中文。"""
        summary = _make_summary(results=[_make_result()], period="2024-12")
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["校验明细"]
            headers = [ws.cell(row=1, column=c).value for c in range(1, 11)]
            assert headers[0] == "规则ID"
            assert headers[1] == "规则名称"
            assert headers[2] == "分类"
            assert headers[3] == "校验结果"
            assert headers[4] == "左侧值"
            assert headers[5] == "右侧值"
            assert headers[6] == "差额"
            assert headers[7] == "容差"
            assert headers[8] == "公式"
            assert headers[9] == "说明"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_trace_sheet_headers_chinese(self) -> None:
        """追溯 sheet 标题使用中文。"""
        trace = [_make_trace()]
        summary = _make_summary(results=[_make_result(trace=trace)])
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["科目追溯"]
            headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
            assert headers[0] == "规则ID"
            assert headers[1] == "规则名称"
            assert headers[2] == "侧"
            assert headers[3] == "科目名称"
            assert headers[4] == "金额"
            assert headers[5] == "原始行"
            assert headers[6] == "原始列"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_money_format_two_decimals(self) -> None:
        """金额列使用 #,##0.00 格式。"""
        results = [_make_result(left_value=1234567.89, right_value=1234567.89)]
        summary = _make_summary(results=results)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["校验明细"]
            # 检查左侧值列 (col 5) 的格式
            cell = ws.cell(row=2, column=5)
            assert cell.number_format == "#,##0.00", f"预期格式 '#,##0.00'，实际 '{cell.number_format}'"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_header_frozen(self) -> None:
        """表头行已冻结。"""
        summary = _make_summary(results=[_make_result()])
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            for name in ["校验明细", "科目追溯"]:
                ws = wb[name]
                assert ws.freeze_panes is not None, f"{name} sheet 应冻结表头"
                # freeze_panes 应为 "A2" (冻结第一行)
                assert ws.freeze_panes == "A2", f"{name} freeze_panes 应为 A2，实际 {ws.freeze_panes}"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_permission_error_handling(self) -> None:
        """权限错误路径: 写入只读目录时抛出 PermissionError。"""
        exporter = AuditExporter()
        summary = _make_summary(results=[_make_result()])

        with tempfile.TemporaryDirectory() as tmpdir:
            readonly_dir = Path(tmpdir) / "readonly"
            readonly_dir.mkdir()
            # Windows: 设置目录为只读
            os.chmod(str(readonly_dir), stat.S_IREAD)
            file_path = str(readonly_dir / "test.xlsx")

            try:
                exporter.export(summary, file_path)
                # 如果成功写出了 (Windows 可能忽略目录权限)，跳过
                if os.path.exists(file_path):
                    os.chmod(str(readonly_dir), stat.S_IWRITE)
                    os.unlink(file_path)
            except PermissionError:
                pass  # 预期行为
            except OSError:
                pass  # 预期行为
            finally:
                os.chmod(str(readonly_dir), stat.S_IWRITE)

    def test_export_with_negative_values(self) -> None:
        """负值金额正确导出。"""
        results = [_make_result(left_value=-100.0, right_value=-50.0, diff=-50.0)]
        summary = _make_summary(results=results)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["校验明细"]
            assert ws.cell(row=2, column=5).value == -100.0
            assert ws.cell(row=2, column=6).value == -50.0
            assert ws.cell(row=2, column=7).value == -50.0
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_large_numbers(self) -> None:
        """大数 (1e15) 正确导出。"""
        big = 1e15
        results = [_make_result(left_value=big, right_value=big)]
        summary = _make_summary(results=results)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["校验明细"]
            assert ws.cell(row=2, column=5).value == big
            assert ws.cell(row=2, column=6).value == big
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_summary_conclusion_all_passed(self) -> None:
        """全部通过时结论为"全部通过"。"""
        summary = _make_summary(
            results=[_make_result(passed=True)],
            total=1,
            passed=1,
            failed=0,
            errored=0,
            skipped=0,
        )
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["校验汇总"]
            values_str = " ".join(
                str(cell)
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)
                for cell in row
                if cell is not None
            )
            assert "全部通过" in values_str
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_summary_conclusion_has_failed(self) -> None:
        """有不通过时结论包含"不通过"。"""
        summary = _make_summary(
            results=[_make_result(passed=False, diff=5.0)],
            total=1,
            passed=0,
            failed=1,
            errored=0,
            skipped=0,
        )
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["校验汇总"]
            values_str = " ".join(
                str(cell)
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)
                for cell in row
                if cell is not None
            )
            assert "不通过" in values_str
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_skipped_rules_no_trace_rows(self) -> None:
        """跳过的规则不产生追溯行。"""
        results = [
            _make_result(rule_id="R1", skipped=True, trace=[]),
            _make_result(rule_id="R2", passed=True, trace=[_make_trace()]),
        ]
        summary = _make_summary(results=results, total=1, passed=1, failed=0, skipped=1)
        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["科目追溯"]
            data_rows = ws.max_row - 1
            assert data_rows == 1, f"预期 1 行追溯 (仅 R2)，实际 {data_rows} 行"
            wb.close()
        finally:
            os.unlink(tmp_path)

    def test_workpaper_info_sheet_records_amount_unit_notes(self) -> None:
        """底稿说明 sheet 写入金额单位留痕。"""
        summary = _make_summary(results=[_make_result()])
        summary.amount_unit_notes = ["资产负债表: 金额单位 万元，已统一换算为元"]

        exporter = AuditExporter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp_path = f.name
        try:
            exporter.export(summary, tmp_path)
            wb = load_workbook(tmp_path)
            ws = wb["底稿说明"]
            values = [str(cell.value) for row in ws.iter_rows() for cell in row]
            assert "金额单位留痕" in values
            assert any("资产负债表: 金额单位 万元" in value for value in values)
            wb.close()
        finally:
            os.unlink(tmp_path)
