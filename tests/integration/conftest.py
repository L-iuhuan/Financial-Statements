"""端到端集成测试 fixtures: 生成真实财务场景 Excel 文件。

使用 openpyxl 在临时目录生成包含三大报表的 Excel 文件,
覆盖资产负债表/利润表/现金流量表的核心科目变量。

场景: 示例科技有限公司 2024 年 12 月财报
- 资产总计 9,000,000 = 负债 4,000,000 + 权益 5,000,000 (BS-BAL-001 ✓)
- 流动+非流动资产 = 资产总计 (BS-BAL-002 ✓)
- 流动+非流动负债 = 负债合计 (BS-BAL-003 ✓)
- 权益各组成部分 = 权益合计 (BS-BAL-004 ✓)
- 营业利润 = 收入 - 成本 - 费用 + 其他收益 (IS 内部 ✓)
- 净利润 = 利润总额 - 所得税 (IS 内部 ✓)
- 现金净增加 = 经营 + 投资 + 筹资 (CF 内部 ✓)
- 期末现金 = 期初 + 净增加 (CF ✓)
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl import Workbook

from fsa.core.engine.registry import RuleRegistry
from fsa.core.importer.importer import ImportService
from fsa.services.validation_service import ValidationService


def _write_balance_sheet(ws) -> None:
    """写入资产负债表数据。"""
    ws.cell(row=1, column=1, value="项目")
    ws.cell(row=1, column=2, value="期末余额")
    ws.cell(row=1, column=3, value="期初余额")

    rows = [
        ("货币资金", 2_000_000, 1_700_000),
        ("应收账款", 1_500_000, 1_300_000),
        ("预付款项", 300_000, 250_000),
        ("其他应收款", 200_000, 180_000),
        ("流动资产合计", 5_000_000, 4_500_000),
        ("在建工程", 1_000_000, 800_000),
        ("非流动资产合计", 4_000_000, 3_500_000),
        ("资产总计", 9_000_000, 8_000_000),
        ("应付账款", 1_200_000, 1_000_000),
        ("预收款项", 800_000, 700_000),
        ("流动负债合计", 3_000_000, 2_700_000),
        ("非流动负债合计", 1_000_000, 1_000_000),
        ("负债合计", 4_000_000, 3_700_000),
        ("实收资本", 3_000_000, 3_000_000),
        ("资本公积", 500_000, 500_000),
        ("其他综合收益", 0, 0),
        ("盈余公积", 200_000, 150_000),
        ("未分配利润", 1_300_000, 650_000),
        ("库存股", 0, 0),
        ("所有者权益合计", 5_000_000, 4_300_000),
    ]

    for idx, (name, ending, beginning) in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=ending)
        ws.cell(row=idx, column=3, value=beginning)


def _write_income_statement(ws) -> None:
    """写入利润表数据。"""
    ws.cell(row=1, column=1, value="项目")
    ws.cell(row=1, column=2, value="本期金额")
    ws.cell(row=1, column=3, value="上期金额")

    rows = [
        ("营业收入", 5_000_000, 4_500_000),
        ("营业成本", 3_500_000, 3_200_000),
        ("税金及附加", 50_000, 45_000),
        ("销售费用", 200_000, 180_000),
        ("管理费用", 300_000, 280_000),
        ("研发费用", 150_000, 120_000),
        ("财务费用", 100_000, 90_000),
        ("资产减值损失", 0, 0),
        ("信用减值损失", 0, 0),
        ("其他收益", 30_000, 25_000),
        ("投资收益", 50_000, 40_000),
        ("公允价值变动收益", 0, 0),
        ("资产处置收益", 0, 0),
        ("营业利润", 780_000, 630_000),
        ("营业外收入", 20_000, 15_000),
        ("营业外支出", 0, 0),
        ("利润总额", 800_000, 645_000),
        ("所得税费用", 200_000, 161_250),
        ("净利润", 600_000, 483_750),
    ]

    for idx, (name, current, prior) in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=current)
        ws.cell(row=idx, column=3, value=prior)


def _write_cash_flow(ws) -> None:
    """写入现金流量表数据。"""
    ws.cell(row=1, column=1, value="项目")
    ws.cell(row=1, column=2, value="本期金额")
    ws.cell(row=1, column=3, value="上期金额")

    rows = [
        ("销售商品、提供劳务收到的现金", 4_800_000, 4_300_000),
        ("经营活动产生的现金流量净额", 800_000, 700_000),
        ("投资活动产生的现金流量净额", -300_000, -250_000),
        ("筹资活动产生的现金流量净额", -200_000, -150_000),
        ("现金及现金等价物净增加额", 300_000, 300_000),
        ("汇率变动对现金的影响", 0, 0),
        ("期初现金及现金等价物余额", 1_700_000, 1_400_000),
        ("期末现金及现金等价物余额", 2_000_000, 1_700_000),
    ]

    for idx, (name, current, prior) in enumerate(rows, start=2):
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=current)
        ws.cell(row=idx, column=3, value=prior)


@pytest.fixture
def sample_excel_path(tmp_path: Path) -> Path:
    """生成真实财务场景 Excel 文件，返回路径。"""
    file_path = tmp_path / "sample_financial_statements.xlsx"
    wb = Workbook()

    ws_bs = wb.active
    if ws_bs is not None:
        ws_bs.title = "资产负债表"
        _write_balance_sheet(ws_bs)

    ws_is = wb.create_sheet("利润表")
    _write_income_statement(ws_is)

    ws_cf = wb.create_sheet("现金流量表")
    _write_cash_flow(ws_cf)

    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture
def sample_registry() -> RuleRegistry:
    """从 CAS 规则库 JSON 加载注册表。"""
    return RuleRegistry.from_json("cas_gouji_rule_library.json")


@pytest.fixture
def imported_reports(sample_excel_path: Path):
    """导入示例 Excel 并返回 Report 列表。"""
    importer = ImportService(period="2024-12")
    return importer.import_file(str(sample_excel_path))


@pytest.fixture
def validation_summary(imported_reports, sample_registry: RuleRegistry):
    """对导入的报表执行校验，返回 ValidationSummary。"""
    service = ValidationService(sample_registry)
    return service.validate(imported_reports, "2024-12")
